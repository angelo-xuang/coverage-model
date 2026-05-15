#!/usr/bin/env python3
"""
NVDA 估值模型构建
用法: python build_nvda_model.py [--input nvda_model_data.json]

基于 coverage-model Skill 的框架，独立生成 NVDA 分部预测 Excel。
不对 Skill 做任何修改。
"""

import sys, json, argparse
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ========== 颜色 ==========
BLUE_FONT = Font(color="0000FF")
BLACK_FONT = Font(color="000000")
GREEN_FONT = Font(color="008000")
GRAY_FONT = Font(color="808080", italic=True)
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BLACK_BOLD = Font(color="000000", bold=True)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
COL_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
OUTPUT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
INPUT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOGGLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NOTE_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
SEGMENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
RIGHT_ALIGN = Alignment(horizontal='right')
WRAP_ALIGN = Alignment(horizontal='left', wrap_text=True, vertical='top')

def sh(ws, row, sc, ec, fill=HEADER_FILL, font=WHITE_BOLD):
    for c in range(sc, ec + 1):
        cl = ws.cell(row=row, column=c)
        cl.fill = fill; cl.font = font
        cl.alignment = Alignment(horizontal='center'); cl.border = THIN_BORDER

def sd(ws, row, col, font=BLACK_FONT, fill=None, nf=None):
    cl = ws.cell(row=row, column=col)
    cl.font = font
    if fill: cl.fill = fill
    cl.border = THIN_BORDER; cl.alignment = RIGHT_ALIGN
    if nf: cl.number_format = nf

def sec(ws, row, label, cols):
    ws.cell(row=row, column=1, value=f"--- {label} ---")
    ws.cell(row=row, column=1).font = WHITE_BOLD
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for c in range(2, cols + 2):
        ws.cell(row=row, column=c).fill = HEADER_FILL

# ========== 工作表构建 ==========

def build_fin_hist(ws, data, ccy_sym='$', us='M'):
    ws.title = '财务历史'
    years = sorted(data.keys())
    ws.cell(row=1, column=1, value='指标'); sh(ws, 1, 1, len(years) + 1)
    for i, y in enumerate(years):
        ws.cell(row=1, column=i + 2, value=y)
        sh(ws, 1, i + 2, i + 2, fill=COL_HEADER_FILL, font=BLACK_BOLD)
    rows = [
        (f'收入 ({ccy_sym}{us})', 'revenue', BLUE_FONT, INPUT_FILL, '#,##0'),
        ('收入增速', 'revenue_growth', BLACK_FONT, None, '0.0%'),
        (f'毛利 ({ccy_sym}{us})', 'gross_profit', BLACK_FONT, None, '#,##0'),
        ('毛利率', 'gross_margin', BLACK_FONT, None, '0.0%'),
        (f'研发费用 ({ccy_sym}{us})', 'rd', BLUE_FONT, INPUT_FILL, '#,##0'),
        ('研发费用率', 'rd_ratio', BLACK_FONT, None, '0.0%'),
        (f'EBIT ({ccy_sym}{us})', 'ebit', BLACK_FONT, None, '#,##0'),
        ('EBIT利润率', 'ebit_margin', BLACK_FONT, None, '0.0%'),
        (f'净利润 ({ccy_sym}{us})', 'net_income', BLUE_FONT, INPUT_FILL, '#,##0'),
        ('净利率', 'net_margin', BLACK_FONT, None, '0.0%'),
    ]
    for ri, (lb, k, fnt, fl, fm) in enumerate(rows, start=2):
        ws.cell(row=ri, column=1, value=lb)
        ws.cell(row=ri, column=1).font = BLACK_BOLD; ws.cell(row=ri, column=1).border = THIN_BORDER
        for ci, y in enumerate(years, start=2):
            v = data[y].get(k)
            if v is not None: ws.cell(row=ri, column=ci, value=v)
            sd(ws, ri, ci, font=fnt, fill=fl, nf=fm)
    ws.column_dimensions['A'].width = 22
    for c in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14


def build_seg_hist(ws, seg_data, ccy_sym='$'):
    ws.title = '分部历史'
    years = sorted(seg_data.keys())
    segs = ['dc', 'gaming', 'proviz', 'auto', 'oem']
    names = {'dc': 'Data Center', 'gaming': 'Gaming', 'proviz': 'Pro Visualization',
             'auto': 'Automotive', 'oem': 'OEM & Other'}
    ws.cell(row=1, column=1, value='分部'); sh(ws, 1, 1, len(years) + 1)
    for i, y in enumerate(years):
        ws.cell(row=1, column=i + 2, value=y)
        sh(ws, 1, i + 2, i + 2, fill=COL_HEADER_FILL, font=BLACK_BOLD)
    for ri, seg in enumerate(segs, start=2):
        ws.cell(row=ri, column=1, value=names[seg])
        ws.cell(row=ri, column=1).font = BLACK_BOLD; ws.cell(row=ri, column=1).border = THIN_BORDER
        for ci, y in enumerate(years, start=2):
            v = seg_data[y].get(seg)
            if v: ws.cell(row=ri, column=ci, value=v)
            sd(ws, ri, ci, font=BLUE_FONT, fill=INPUT_FILL, nf='#,##0.0')
    # 占比行
    for ri, seg in enumerate(segs, start=7):
        ws.cell(row=ri, column=1, value=f'{names[seg]} 占比')
        ws.cell(row=ri, column=1).font = BLACK_BOLD; ws.cell(row=ri, column=1).border = THIN_BORDER
        for ci, y in enumerate(years, start=2):
            c = get_column_letter(ci)
            row_num = 2 + ['dc', 'gaming', 'proviz', 'auto', 'oem'].index(seg)
            total_row = 7
            ws.cell(row=ri, column=ci, value=f'={c}{row_num}/{c}{total_row}')
            sd(ws, ri, ci, font=BLACK_FONT, nf='0.0%')
    ws.column_dimensions['A'].width = 22
    for c in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14


def build_assumptions(ws, assumptions, sources, ccy_sym='$'):
    ws.title = '假设'
    sn = '假设'

    ws.cell(row=1, column=1, value='活跃情景')
    ws.cell(row=1, column=1).font = WHITE_BOLD; ws.cell(row=1, column=1).fill = HEADER_FILL; ws.cell(row=1, column=1).border = THIN_BORDER
    ws.cell(row=1, column=2, value=2)
    ws.cell(row=1, column=2).font = BLACK_BOLD; ws.cell(row=1, column=2).fill = TOGGLE_FILL; ws.cell(row=1, column=2).border = THIN_BORDER
    dv = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    dv.error = "1=Bear, 2=Base, 3=Bull"
    ws.add_data_validation(dv); dv.add(ws["B1"])

    ws.cell(row=2, column=1, value='指标'); sh(ws, 2, 1, 7)
    for c, t in [(3, 'Bear'), (4, 'Base'), (5, 'Bull')]:
        ws.cell(row=2, column=c, value=t)
        sh(ws, 2, c, c, fill=COL_HEADER_FILL, font=BLACK_BOLD)
    ws.cell(row=2, column=6, value='活跃值'); sh(ws, 2, 6, 6, fill=OUTPUT_FILL, font=BLACK_BOLD)
    ws.cell(row=2, column=7, value='来源/推导依据')
    sh(ws, 2, 7, 7, fill=PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"), font=Font(color="666666", bold=True))

    rows_def = [
        ('', f'--- Data Center ---', None, None, None, True),
        ('dc', f'AIDC CapEx 起始 ({ccy_sym}B)', 'tam_start', '#,##0.0', 'tam_start'),
        ('dc', 'AIDC CapEx CAGR Y1-3', 'tam_cagr_1_3', '0.0%', 'tam_cagr_1_3'),
        ('dc', 'AIDC CapEx CAGR Y4-6', 'tam_cagr_4_6', '0.0%', 'tam_cagr_4_6'),
        ('dc', 'AIDC CapEx CAGR Y7-10', 'tam_cagr_7_10', '0.0%', 'tam_cagr_7_10'),
        ('dc', 'NVDA可触达率', 'reachable_rate', '0.0%', 'reachable_rate'),
        ('dc', 'AI GPU市占率 2025', 'share_start', '0.0%', 'share_start'),
        ('dc', 'AI GPU市占率 2035', 'share_end', '0.0%', 'share_end'),
        ('', f'--- Gaming ---', None, None, None, True),
        ('g', f'Gaming起始收入 ({ccy_sym}B)', 'gaming_rev', '#,##0.0', 'gaming_rev'),
        ('g', 'Gaming CAGR Y1-3', 'gaming_cagr_1_3', '0.0%', 'gaming_cagr_1_3'),
        ('g', 'Gaming CAGR Y4-6', 'gaming_cagr_4_6', '0.0%', 'gaming_cagr_4_6'),
        ('g', 'Gaming CAGR Y7-10', 'gaming_cagr_7_10', '0.0%', 'gaming_cagr_7_10'),
        ('', f'--- Pro Visualization ---', None, None, None, True),
        ('pv', f'ProViz起始收入 ({ccy_sym}B)', 'proviz_rev', '#,##0.0', 'proviz_rev'),
        ('pv', 'ProViz CAGR Y1-3', 'proviz_cagr_1_3', '0.0%', 'proviz_cagr_1_3'),
        ('pv', 'ProViz CAGR Y4-6', 'proviz_cagr_4_6', '0.0%', 'proviz_cagr_4_6'),
        ('pv', 'ProViz CAGR Y7-10', 'proviz_cagr_7_10', '0.0%', 'proviz_cagr_7_10'),
        ('', f'--- Automotive ---', None, None, None, True),
        ('a', f'Auto起始收入 ({ccy_sym}B)', 'auto_rev', '#,##0.0', 'auto_rev'),
        ('a', 'Auto CAGR Y1-3', 'auto_cagr_1_3', '0.0%', 'auto_cagr_1_3'),
        ('a', 'Auto CAGR Y4-6', 'auto_cagr_4_6', '0.0%', 'auto_cagr_4_6'),
        ('a', 'Auto CAGR Y7-10', 'auto_cagr_7_10', '0.0%', 'auto_cagr_7_10'),
        ('', f'--- OEM & Other ---', None, None, None, True),
        ('o', f'OEM起始收入 ({ccy_sym}B)', 'oem_rev', '#,##0.0', 'oem_rev'),
        ('', f'--- 利润率 ---', None, None, None, True),
        ('', '起始毛利率', 'gm_start', '0.0%', 'gm_start'),
        ('', '终年毛利率', 'gm_end', '0.0%', 'gm_end'),
        ('', '起始费用率', 'expense_start', '0.0%', 'expense_start'),
        ('', '终年费用率', 'expense_end', '0.0%', 'expense_end'),
        ('', '税率', 'tax_rate', '0.0%', 'tax_rate'),
        ('', f'--- 估值 ---', None, None, None, True),
        ('', '起始PE', 'pe_start', '0.0', 'pe_start'),
        ('', '终年PE', 'pe_end', '0.0', 'pe_end'),
        ('', f'股份数 (M)', 'shares_m', '#,##0', 'shares_m'),
    ]

    row = 3
    row_map = {}
    for entry in rows_def:
        seg, label, key, fmt = entry[0], entry[1], entry[2], entry[3]
        sk = entry[4] if len(entry) > 4 else None
        is_header = entry[5] if len(entry) > 5 else False

        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = BLACK_BOLD; ws.cell(row=row, column=1).border = THIN_BORDER

        if is_header or key is None:
            ws.cell(row=row, column=1).fill = SEGMENT_FILL
            for c in range(2, 8):
                ws.cell(row=row, column=c).fill = SEGMENT_FILL; ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1; continue

        for ci, sc in enumerate(['bear', 'base', 'bull'], start=3):
            v = assumptions.get(sc, {}).get(key)
            if v is not None: ws.cell(row=row, column=ci, value=v)
            sd(ws, row, ci, font=BLUE_FONT, fill=INPUT_FILL, nf=fmt)

        ws.cell(row=row, column=6, value=f'=INDEX(C{row}:E{row},$B$1)')
        sd(ws, row, 6, font=GREEN_FONT, fill=OUTPUT_FILL, nf=fmt)

        st = sources.get(key, sources.get(sk, ''))
        ws.cell(row=row, column=7, value=st)
        ws.cell(row=row, column=7).font = GRAY_FONT; ws.cell(row=row, column=7).fill = NOTE_FILL
        ws.cell(row=row, column=7).border = THIN_BORDER; ws.cell(row=row, column=7).alignment = WRAP_ALIGN

        row_map[key] = row
        row += 1

    ws.column_dimensions['A'].width = 28
    for c in ['B', 'C', 'D', 'E', 'F']: ws.column_dimensions[c].width = 14
    ws.column_dimensions['G'].width = 55
    return row_map


def build_forecast(ws, assumptions, hist_data, row_map, ccy_sym='$', pyears=10, current_price=None, current_pe=None):
    ws.title = '预测'
    sn = '假设'
    VAL_PRICE = current_price or 100

    # 获取 last actual year
    hist_years = sorted(hist_data.keys())
    ly_str = hist_years[-1]
    ly = int(ly_str.replace('A', ''))
    yrs = [ly_str] + [f"{ly + i + 1}E" for i in range(pyears)]
    TC = len(yrs)

    def asum(key):
        r = row_map[key]
        return f"'{sn}'!$F${r}"

    ws.cell(row=1, column=1, value='指标')
    for i, y in enumerate(yrs):
        ws.cell(row=1, column=i + 2, value=y)
    sh(ws, 1, 1, TC + 1)
    sh(ws, 1, 2, TC + 1, fill=COL_HEADER_FILL, font=BLACK_BOLD)

    # 历史数据
    hist_rev = hist_data.get(ly_str, {}).get('revenue', 0)
    hist_gm = hist_data.get(ly_str, {}).get('gross_margin', 0)
    hist_ni = hist_data.get(ly_str, {}).get('net_income', 0)
    hist_growth = hist_data.get(ly_str, {}).get('revenue_growth', 0)

    # ---------- DC TAM 框架 ----------
    sec(ws, 2, '收入推导: Data Center', TC)
    # R3-R6: TAM params
    for i, (label, key, fmt) in enumerate([
        (f'AIDC CapEx 起始 ({ccy_sym}B)', 'tam_start', '#,##0.0'),
        ('AIDC CapEx CAGR Y1-3', 'tam_cagr_1_3', '0.0%'),
        ('AIDC CapEx CAGR Y4-6', 'tam_cagr_4_6', '0.0%'),
        ('AIDC CapEx CAGR Y7-10', 'tam_cagr_7_10', '0.0%'),
    ], start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = BLACK_BOLD; ws.cell(row=i, column=1).border = THIN_BORDER
        for j in range(TC):
            ws.cell(row=i, column=j + 2, value=f"={asum(key)}")
            sd(ws, i, j + 2, font=GREEN_FONT, nf=fmt)

    # R7: AIDC CapEx (计算)
    r7 = 7
    ws.cell(row=r7, column=1, value=f'AIDC CapEx ({ccy_sym}B)')
    ws.cell(row=r7, column=1).font = BLACK_BOLD; ws.cell(row=r7, column=1).border = THIN_BORDER
    ws.cell(row=r7, column=2, value=f"={asum('tam_start')}")
    sd(ws, r7, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='#,##0.0')
    for i in range(1, TC):
        col = i + 2
        c = get_column_letter(col)
        pc = get_column_letter(col - 1)
        if i <= 2: cr = 4
        elif i <= 5: cr = 5
        else: cr = 6
        if i == 1:
            ws.cell(row=r7, column=col, value=f"={c}3*(1+{c}{cr})")
        else:
            ws.cell(row=r7, column=col, value=f"={pc}{r7}*(1+{c}{cr})")
        sd(ws, r7, col, font=BLACK_FONT, nf='#,##0.0')

    # R8: 可触达率
    ws.cell(row=8, column=1, value='NVDA可触达率')
    ws.cell(row=8, column=1).font = BLACK_BOLD; ws.cell(row=8, column=1).border = THIN_BORDER
    for j in range(TC):
        ws.cell(row=8, column=j + 2, value=f"={asum('reachable_rate')}")
        sd(ws, 8, j + 2, font=GREEN_FONT, nf='0.0%')

    # R9: 可触达市场
    r9 = 9
    ws.cell(row=r9, column=1, value=f'可触达市场 ({ccy_sym}B)')
    ws.cell(row=r9, column=1).font = BLACK_BOLD; ws.cell(row=r9, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r9, column=j + 2, value=f"={c}7*{c}8")
        sd(ws, r9, j + 2, font=BLACK_FONT, nf='#,##0.0')

    # R10-R11: 市占率 params
    for i, (label, key) in enumerate([('AI GPU市占率起始', 'share_start'), ('AI GPU市占率终年', 'share_end')], start=10):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = BLACK_BOLD; ws.cell(row=i, column=1).border = THIN_BORDER
        for j in range(TC):
            ws.cell(row=i, column=j + 2, value=f"={asum(key)}")
            sd(ws, i, j + 2, font=GREEN_FONT, nf='0.0%')

    # R12: 市占率 (线性插值)
    r12 = 12
    ws.cell(row=r12, column=1, value='市占率')
    ws.cell(row=r12, column=1).font = BLACK_BOLD; ws.cell(row=r12, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        if j == 0:
            ws.cell(row=r12, column=2, value=f"={c}10")
        else:
            ws.cell(row=r12, column=j + 2, value=f"={c}10+({c}11-{c}10)*{j}/{pyears}")
        sd(ws, r12, j + 2, font=BLACK_FONT, nf='0.0%')

    # R13: DC收入
    r13 = 13
    ws.cell(row=r13, column=1, value=f'DC收入 ({ccy_sym}B)')
    ws.cell(row=r13, column=1).font = BLACK_BOLD; ws.cell(row=r13, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r13, column=j + 2, value=f"={c}9*{c}12")
        sd(ws, r13, j + 2, font=BLACK_FONT, fill=OUTPUT_FILL, nf='#,##0.0')

    # ---------- 其他分部 ----------
    # 每个分部: 起始收入 × (1+CAGR各段) 复合
    segments = [
        ('Gaming', 15, 'gaming_rev', 'gaming_cagr_1_3', 'gaming_cagr_4_6', 'gaming_cagr_7_10'),
        ('Pro Visualization', 17, 'proviz_rev', 'proviz_cagr_1_3', 'proviz_cagr_4_6', 'proviz_cagr_7_10'),
        ('Automotive', 19, 'auto_rev', 'auto_cagr_1_3', 'auto_cagr_4_6', 'auto_cagr_7_10'),
        ('OEM & Other', 21, 'oem_rev', None, None, None),  # OEM 固定
    ]

    for seg_name, hdr_row, rev_key, c13k, c46k, c710k in segments:
        # header
        ws.cell(row=hdr_row, column=1, value=f'-- {seg_name} --')
        ws.cell(row=hdr_row, column=1).font = Font(bold=True, color="1F4E79")
        ws.cell(row=hdr_row, column=1).border = THIN_BORDER

        rev_row = hdr_row + 1
        ws.cell(row=rev_row, column=1, value=f'{seg_name}收入 ({ccy_sym}B)')
        ws.cell(row=rev_row, column=1).font = BLACK_BOLD; ws.cell(row=rev_row, column=1).border = THIN_BORDER
        for j in range(TC):
            col = j + 2
            pc = get_column_letter(col - 1) if col > 2 else ''
            if c13k is None:  # OEM: 固定值
                ws.cell(row=rev_row, column=col, value=f"={asum(rev_key)}")
                sd(ws, rev_row, col, font=GREEN_FONT, nf='#,##0.0')
            else:
                if j == 0:
                    ws.cell(row=rev_row, column=col, value=f"={asum(rev_key)}")
                    sd(ws, rev_row, col, font=BLUE_FONT, fill=INPUT_FILL, nf='#,##0.0')
                else:
                    c = get_column_letter(col)
                    pc = get_column_letter(col - 1)
                    if (j - 1 + 1) <= 3: cagr_k = c13k
                    elif (j - 1 + 1) <= 6: cagr_k = c46k
                    else: cagr_k = c710k
                    ws.cell(row=rev_row, column=col, value=f"={pc}{rev_row}*(1+{asum(cagr_k)})")
                    sd(ws, rev_row, col, font=BLACK_FONT, nf='#,##0.0')

    # ---------- 汇总 ----------
    sec(ws, 23, '总收入汇总', TC)
    r23 = 24
    ws.cell(row=r23, column=1, value=f'总收入 ({ccy_sym}B)')
    ws.cell(row=r23, column=1).font = BLACK_BOLD; ws.cell(row=r23, column=1).border = THIN_BORDER
    if hist_rev:
        ws.cell(row=r23, column=2, value=hist_rev)
        sd(ws, r23, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='#,##0')
    for j in range(1, TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r23, column=j + 2, value=f"={c}13+{c}16+{c}18+{c}20+{c}22")
        sd(ws, r23, j + 2, font=BLACK_FONT, fill=OUTPUT_FILL, nf='#,##0.0')

    r24 = 25
    ws.cell(row=r24, column=1, value='总收入增速')
    ws.cell(row=r24, column=1).font = BLACK_BOLD; ws.cell(row=r24, column=1).border = THIN_BORDER
    ws.cell(row=r24, column=2, value=hist_growth if hist_growth else None)
    sd(ws, r24, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='0.0%')
    c3 = get_column_letter(3)
    if hist_rev:
        ws.cell(row=r24, column=3, value=f"=({c3}24*1000-{hist_rev})/{hist_rev}")
    sd(ws, r24, 3, font=BLACK_FONT, nf='0.0%')
    for j in range(2, TC):
        c = get_column_letter(j + 2); pc = get_column_letter(j + 1)
        ws.cell(row=r24, column=j + 2, value=f"=({c}24-{pc}24)/{pc}24")
        sd(ws, r24, j + 2, font=BLACK_FONT, nf='0.0%')

    # ---------- 利润 ----------
    sec(ws, 27, '利润推导', TC)
    # R28: 毛利率起始, R29: 毛利率终年
    for i, (label, key) in enumerate([('起始毛利率', 'gm_start'), ('终年毛利率', 'gm_end')], start=28):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = BLACK_BOLD; ws.cell(row=i, column=1).border = THIN_BORDER
        for j in range(TC):
            ws.cell(row=i, column=j + 2, value=f"={asum(key)}")
            sd(ws, i, j + 2, font=GREEN_FONT, nf='0.0%')

    # R30: 毛利率 (线性)
    r30 = 30
    ws.cell(row=r30, column=1, value='毛利率')
    ws.cell(row=r30, column=1).font = BLACK_BOLD; ws.cell(row=r30, column=1).border = THIN_BORDER
    if hist_gm:
        ws.cell(row=r30, column=2, value=hist_gm)
        sd(ws, r30, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='0.0%')
    for j in range(1, TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r30, column=j + 2, value=f"={c}28+({c}29-{c}28)*{j - 1}/{pyears - 1}")
        sd(ws, r30, j + 2, font=BLACK_FONT, nf='0.0%')

    # R31: 毛利 = 总收入 × 毛利率
    r31 = 31
    ws.cell(row=r31, column=1, value=f'毛利 ({ccy_sym}B)')
    ws.cell(row=r31, column=1).font = BLACK_BOLD; ws.cell(row=r31, column=1).border = THIN_BORDER
    if hist_rev and hist_gm:
        ws.cell(row=r31, column=2, value=hist_rev * hist_gm)
        sd(ws, r31, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='#,##0')
    for j in range(1, TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r31, column=j + 2, value=f"={c}24*{c}30")
        sd(ws, r31, j + 2, font=BLACK_FONT, nf='#,##0.0')

    # R32: 费用率起始, R33: 费用率终年
    for i, (label, key) in enumerate([('起始费用率', 'expense_start'), ('终年费用率', 'expense_end')], start=32):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = BLACK_BOLD; ws.cell(row=i, column=1).border = THIN_BORDER
        for j in range(TC):
            ws.cell(row=i, column=j + 2, value=f"={asum(key)}")
            sd(ws, i, j + 2, font=GREEN_FONT, nf='0.0%')

    # R34: 费用率 (线性)
    r34 = 34
    ws.cell(row=r34, column=1, value='费用率')
    ws.cell(row=r34, column=1).font = BLACK_BOLD; ws.cell(row=r34, column=1).border = THIN_BORDER
    # B列: 历史费用率 = 毛利率 - EBIT率
    hist_ebit_m = hist_data.get(ly_str, {}).get('ebit_margin', 0)
    hist_exp_ratio = hist_gm - hist_ebit_m if hist_gm and hist_ebit_m else None
    if hist_exp_ratio:
        ws.cell(row=r34, column=2, value=hist_exp_ratio)
        sd(ws, r34, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='0.0%')
    for j in range(1, TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r34, column=j + 2, value=f"={c}32+({c}33-{c}32)*{j - 1}/{pyears - 1}")
        sd(ws, r34, j + 2, font=BLACK_FONT, nf='0.0%')

    # R35: 运营费用
    r35 = 35
    ws.cell(row=r35, column=1, value=f'运营费用 ({ccy_sym}B)')
    ws.cell(row=r35, column=1).font = BLACK_BOLD; ws.cell(row=r35, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r35, column=j + 2, value=f"={c}24*{c}34")
        sd(ws, r35, j + 2, font=BLACK_FONT, nf='#,##0.0')

    # R36: EBIT
    r36 = 36
    ws.cell(row=r36, column=1, value=f'EBIT ({ccy_sym}B)')
    ws.cell(row=r36, column=1).font = BLACK_BOLD; ws.cell(row=r36, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r36, column=j + 2, value=f"={c}31-{c}35")
        sd(ws, r36, j + 2, font=BLACK_FONT, nf='#,##0.0')

    # R37: 税率
    ws.cell(row=37, column=1, value='税率')
    ws.cell(row=37, column=1).font = BLACK_BOLD; ws.cell(row=37, column=1).border = THIN_BORDER
    for j in range(TC):
        ws.cell(row=37, column=j + 2, value=f"={asum('tax_rate')}")
        sd(ws, 37, j + 2, font=GREEN_FONT, nf='0.0%')

    # R38: 净利润
    r38 = 38
    ws.cell(row=r38, column=1, value=f'净利润 ({ccy_sym}B)')
    ws.cell(row=r38, column=1).font = BLACK_BOLD; ws.cell(row=r38, column=1).border = THIN_BORDER
    if hist_ni:
        ws.cell(row=r38, column=2, value=hist_ni)
        sd(ws, r38, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='#,##0')
    for j in range(1, TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r38, column=j + 2, value=f"={c}36*(1-{c}37)")
        sd(ws, r38, j + 2, font=BLACK_FONT, fill=OUTPUT_FILL, nf='#,##0.0')

    # ---------- 估值 ----------
    sec(ws, 40, '估值', TC)
    for i, (label, key) in enumerate([('起始PE', 'pe_start'), ('终年PE', 'pe_end')], start=41):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = BLACK_BOLD; ws.cell(row=i, column=1).border = THIN_BORDER
        for j in range(TC):
            ws.cell(row=i, column=j + 2, value=f"={asum(key)}")
            sd(ws, i, j + 2, font=GREEN_FONT, nf='0.0')

    # R43: 目标PE (线性)
    r43 = 43
    ws.cell(row=r43, column=1, value='目标PE')
    ws.cell(row=r43, column=1).font = BLACK_BOLD; ws.cell(row=r43, column=1).border = THIN_BORDER
    if current_pe:
        ws.cell(row=r43, column=2, value=current_pe)
        sd(ws, r43, 2, font=BLUE_FONT, fill=INPUT_FILL, nf='0.0')
    for j in range(1, TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r43, column=j + 2, value=f"={c}41+({c}42-{c}41)*{j - 1}/{pyears - 1}")
        sd(ws, r43, j + 2, font=BLACK_FONT, nf='0.0')

    # R44: 隐含市值
    r44 = 44
    ws.cell(row=r44, column=1, value=f'隐含市值 ({ccy_sym}B)')
    ws.cell(row=r44, column=1).font = BLACK_BOLD; ws.cell(row=r44, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r44, column=j + 2, value=f"={c}38*{c}43")
        sd(ws, r44, j + 2, font=BLACK_FONT, fill=OUTPUT_FILL, nf='#,##0.0')

    # R45: 股份数
    ws.cell(row=45, column=1, value='股份数 (M)')
    ws.cell(row=45, column=1).font = BLACK_BOLD; ws.cell(row=45, column=1).border = THIN_BORDER
    for j in range(TC):
        ws.cell(row=45, column=j + 2, value=f"={asum('shares_m')}")
        sd(ws, 45, j + 2, font=GREEN_FONT, nf='#,##0')

    # R46: 隐含股价
    r46 = 46
    ws.cell(row=r46, column=1, value=f'隐含股价 ({ccy_sym})')
    ws.cell(row=r46, column=1).font = BLACK_BOLD; ws.cell(row=r46, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r46, column=j + 2, value=f"={c}44*1000/{c}45")
        sd(ws, r46, j + 2, font=BLACK_FONT, fill=OUTPUT_FILL, nf='#,##0.00')

    # R47: vs当前溢价
    r47 = 47
    ws.cell(row=r47, column=1, value='vs 当前股价溢价')
    ws.cell(row=r47, column=1).font = BLACK_BOLD; ws.cell(row=r47, column=1).border = THIN_BORDER
    for j in range(TC):
        c = get_column_letter(j + 2)
        ws.cell(row=r47, column=j + 2, value=f"=({c}46-{VAL_PRICE})/{VAL_PRICE}")
        sd(ws, r47, j + 2, font=BLACK_FONT, nf='0.0%')

    ws.column_dimensions['A'].width = 28
    for c in range(2, TC + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14


def main():
    parser = argparse.ArgumentParser(description='NVDA估值模型Excel生成')
    parser.add_argument('--input', default='../../out/nvda_model_data.json', help='模型数据JSON')
    parser.add_argument('--output', default='NVDA_coverage_model.xlsx', help='输出文件名')
    parser.add_argument('--pyears', type=int, default=10)
    args = parser.parse_args()

    with open(args.input, 'r', encoding='utf-8') as f:
        raw = json.load(f)
    hist_data = raw.get('hist_data', {})
    seg_hist = raw.get('segment_history', {})
    val_data = raw.get('val_data', {})
    assumptions = raw.get('assumptions', {})
    sources = raw.get('sources', {})
    ccy_sym = raw.get('currency_symbol', '$')
    us = raw.get('unit_suffix', 'M')

    wb = Workbook()

    ws1 = wb.active
    build_fin_hist(ws1, hist_data, ccy_sym, us)

    if seg_hist:
        ws_seg = wb.create_sheet()
        build_seg_hist(ws_seg, seg_hist, ccy_sym)

    if assumptions:
        ws_asm = wb.create_sheet()
        row_map = build_assumptions(ws_asm, assumptions, sources, ccy_sym)

        ws_fc = wb.create_sheet()
        cp = val_data.get('current_price')
        build_forecast(ws_fc, assumptions, hist_data, row_map, ccy_sym, args.pyears,
                      current_price=cp, current_pe=val_data.get('current_pe'))

    out_path = Path(args.output)
    wb.save(str(out_path))
    print(f'NVDA模型已生成: {out_path}')
    return str(out_path)

if __name__ == '__main__':
    main()