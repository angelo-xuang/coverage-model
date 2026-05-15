#!/usr/bin/env python3
"""
tsla_build_model.py — Tesla (TSLA) Coverage Model
四分部预测：汽车 + 能源 + 服务 + Optimus
输出: ./out/TSLA_coverage_model.xlsx
"""

import sys, json, argparse
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ===== 样式定义 =====
BLUE_FONT = Font(color="0000FF")
BLACK_FONT = Font(color="000000")
GREEN_FONT = Font(color="008000")
GRAY_FONT = Font(color="808080", italic=True)
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BLACK_BOLD = Font(color="000000", bold=True)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
COL_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
OUTPUT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
INPUT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOGGLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NOTE_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
SEGMENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
RIGHT_ALIGN = Alignment(horizontal='right')
WRAP_ALIGN = Alignment(horizontal='left', wrap_text=True, vertical='top')

LABELS = {
    'sheet_hist': '财务历史',
    'sheet_segments': '业务分部',
    'sheet_val': '估值历史',
    'sheet_assumptions': '假设',
    'sheet_forecast': '预测',
    'sheet_sanity': '合理性检验',
    'sheet_fy': '财年信息',
    'active_scenario': '活跃情景',
    'source_label': '来源/推导依据',
    # 分部标签
    'auto': '汽车',
    'energy': '能源存储',
    'services': '服务及其他',
    'optimus': 'Optimus 机器人',
    'total': '总收入',
    # 假设参数标签
    'ev_tam_2025': '全球EV销量 2025 (M辆)',
    'ev_tam_cagr_13': 'EV TAM CAGR Y1-Y3',
    'ev_tam_cagr_46': 'EV TAM CAGR Y4-Y6',
    'ev_tam_cagr_710': 'EV TAM CAGR Y7-Y10',
    'tsla_share_2025': 'Tesla EV市占率 2025',
    'tsla_share_2035': 'Tesla EV市占率 2035',
    'asp_2025': 'ASP 2025 ($K)',
    'asp_2035': 'ASP 2035 ($K)',
    'storage_tam_2025': '全球储能TAM 2025 ($B)',
    'storage_cagr_13': '储能TAM CAGR Y1-Y3',
    'storage_cagr_46': '储能TAM CAGR Y4-Y6',
    'storage_cagr_710': '储能TAM CAGR Y7-Y10',
    'storage_share_2025': 'Tesla储能市占率 2025',
    'storage_share_2035': 'Tesla储能市占率 2035',
    'svc_rev_2025': '服务收入 2025 ($B)',
    'svc_cagr_13': '服务 CAGR Y1-Y3',
    'svc_cagr_46': '服务 CAGR Y4-Y6',
    'svc_cagr_710': '服务 CAGR Y7-Y10',
    'opt_units_2027': 'Optimus 2027出货量 (M)',
    'opt_unit_cagr_2730': 'Optimus 出货 CAGR 2027-30',
    'opt_unit_cagr_3135': 'Optimus 出货 CAGR 2031-35',
    'opt_asp_2027': 'Optimus ASP 2027 ($K)',
    'opt_asp_2035': 'Optimus ASP 2035 ($K)',
    # 通用参数
    'gm_start': '起始毛利率',
    'gm_end': '终年毛利率',
    'exp_start': '起始费用率',
    'exp_end': '终年费用率',
    'tax_rate': '税率',
    'pe_start': '起始 PE',
    'pe_end': '终年 PE',
    'shares_m': '股份数 (M)',
    # 预测表标签
    'fc_revenue': '收入',
    'fc_rev_growth': '收入增速',
    'fc_gm': '毛利率',
    'fc_gross_profit': '毛利',
    'fc_exp_ratio': '费用率',
    'fc_opex': '运营费用',
    'fc_ebit': 'EBIT',
    'fc_tax_rate': '税率',
    'fc_net_income': '净利润',
    'fc_target_pe': '目标 PE',
    'fc_implied_mkt_cap': '隐含市值',
    'fc_shares': '股份数 (M)',
    'fc_implied_price': '隐含股价',
    'fc_premium': 'vs 当前溢价',
}

# ===== 历史数据 =====
HIST_DATA = {
    '2021A': {'revenue': 53.823, 'auto_revenue': 47.2, 'energy_revenue': 2.8, 'service_revenue': 3.8},
    '2022A': {'revenue': 81.462, 'cogs': 60.609, 'gross_profit': 20.853,
              'rd': 3.075, 'sga': 3.946, 'ebit': 13.832, 'net_income': 12.583,
              'gross_margin': 0.256, 'ebit_margin': 0.1698, 'net_margin': 0.1545,
              'rd_ratio': 0.0377, 'sga_ratio': 0.0484,
              'auto_revenue': 71.5, 'energy_revenue': 3.9, 'service_revenue': 6.1},
    '2023A': {'revenue': 96.773, 'cogs': 79.113, 'gross_profit': 17.660,
              'rd': 3.969, 'sga': 4.800, 'ebit': 8.891, 'net_income': 14.999,
              'gross_margin': 0.1825, 'ebit_margin': 0.0919, 'net_margin': 0.1550,
              'rd_ratio': 0.0410, 'sga_ratio': 0.0496, 'revenue_growth': 0.188,
              'auto_revenue': 82.4, 'energy_revenue': 6.0, 'service_revenue': 8.3},
    '2024A': {'revenue': 97.690, 'cogs': 80.240, 'gross_profit': 17.450,
              'rd': 4.540, 'sga': 5.150, 'ebit': 7.760, 'net_income': 7.130,
              'gross_margin': 0.1786, 'ebit_margin': 0.0794, 'net_margin': 0.0730,
              'rd_ratio': 0.0465, 'sga_ratio': 0.0527, 'revenue_growth': 0.0095,
              'auto_revenue': 77.1, 'energy_revenue': 10.1, 'service_revenue': 10.5},
    '2025A': {'revenue': 94.827, 'cogs': 77.733, 'gross_profit': 17.094,
              'rd': 6.411, 'sga': 5.834, 'ebit': 4.849, 'net_income': 3.794,
              'gross_margin': 0.1803, 'ebit_margin': 0.0511, 'net_margin': 0.0400,
              'rd_ratio': 0.0676, 'sga_ratio': 0.0615, 'revenue_growth': -0.0293,
              'auto_revenue': 69.5, 'energy_revenue': 12.8, 'service_revenue': 12.5},
}

VAL_DATA = {
    'current_market_cap': 1628e9,
    'current_price': 433.45,
    'current_pe': 405.1,
    'forward_pe': 172.4,
    'shares_outstanding': 3755723871,
    'fifty_two_week_high': 498.83,
    'fifty_two_week_low': 273.21,
    'pe_5y_low': 34,
    'pe_5y_median': 100,
    'pe_5y_high': 1120,
}

# ===== 假设 =====
ASSUMPTIONS = {
    'bear': {
        'auto': {
            'ev_tam_2025': 17.5,       # M units
            'tam_cagr_1_3': 0.08,
            'tam_cagr_4_6': 0.05,
            'tam_cagr_7_10': 0.03,
            'share_2025': 0.095,
            'share_2035': 0.05,
            'asp_2025': 38.0,          # $K
            'asp_2035': 25.0,
        },
        'energy': {
            'tam_2025': 50.0,          # $B
            'tam_cagr_1_3': 0.20,
            'tam_cagr_4_6': 0.12,
            'tam_cagr_7_10': 0.08,
            'share_2025': 0.25,
            'share_2035': 0.18,
        },
        'services': {
            'rev_2025': 12.5,          # $B
            'cagr_1_3': 0.08,
            'cagr_4_6': 0.05,
            'cagr_7_10': 0.03,
        },
        'optimus': {
            'units_2027': 0.005,       # M = 5000 units
            'unit_cagr_27_30': 0.30,
            'unit_cagr_31_35': 0.20,
            'asp_2027': 30.0,          # $K
            'asp_2035': 20.0,
        },
        'gm_start': 0.17,
        'gm_end': 0.15,
        'exp_start': 0.14,
        'exp_end': 0.13,
        'tax_rate': 0.21,
        'pe_start': 35,
        'pe_end': 12,
        'shares_m': 3756,
    },
    'base': {
        'auto': {
            'ev_tam_2025': 17.5,
            'tam_cagr_1_3': 0.15,
            'tam_cagr_4_6': 0.10,
            'tam_cagr_7_10': 0.06,
            'share_2025': 0.095,
            'share_2035': 0.07,
            'asp_2025': 38.0,
            'asp_2035': 32.0,
        },
        'energy': {
            'tam_2025': 50.0,
            'tam_cagr_1_3': 0.30,
            'tam_cagr_4_6': 0.20,
            'tam_cagr_7_10': 0.12,
            'share_2025': 0.25,
            'share_2035': 0.28,
        },
        'services': {
            'rev_2025': 12.5,
            'cagr_1_3': 0.12,
            'cagr_4_6': 0.08,
            'cagr_7_10': 0.05,
        },
        'optimus': {
            'units_2027': 0.010,       # 10K
            'unit_cagr_27_30': 0.60,
            'unit_cagr_31_35': 0.40,
            'asp_2027': 30.0,
            'asp_2035': 30.0,
        },
        'gm_start': 0.18,
        'gm_end': 0.21,
        'exp_start': 0.15,
        'exp_end': 0.13,
        'tax_rate': 0.18,
        'pe_start': 80,
        'pe_end': 25,
        'shares_m': 3756,
    },
    'bull': {
        'auto': {
            'ev_tam_2025': 17.5,
            'tam_cagr_1_3': 0.20,
            'tam_cagr_4_6': 0.15,
            'tam_cagr_7_10': 0.10,
            'share_2025': 0.095,
            'share_2035': 0.10,
            'asp_2025': 38.0,
            'asp_2035': 42.0,
        },
        'energy': {
            'tam_2025': 50.0,
            'tam_cagr_1_3': 0.40,
            'tam_cagr_4_6': 0.28,
            'tam_cagr_7_10': 0.18,
            'share_2025': 0.25,
            'share_2035': 0.35,
        },
        'services': {
            'rev_2025': 12.5,
            'cagr_1_3': 0.15,
            'cagr_4_6': 0.12,
            'cagr_7_10': 0.08,
        },
        'optimus': {
            'units_2027': 0.050,       # 50K
            'unit_cagr_27_30': 1.00,
            'unit_cagr_31_35': 0.60,
            'asp_2027': 30.0,
            'asp_2035': 50.0,
        },
        'gm_start': 0.20,
        'gm_end': 0.26,
        'exp_start': 0.13,
        'exp_end': 0.10,
        'tax_rate': 0.15,
        'pe_start': 120,
        'pe_end': 35,
        'shares_m': 3756,
    },
}

SOURCES = {
    'ev_tam_2025': "2025全球EV销量 ~1750万辆 (IEA/Counterpoint)\n历史: 2021 6.6M→2025 17.5M",
    'ev_tam_cagr': "Base Y1-3 15%=IEA中期预测; 增速递减反映渗透率饱和\nBull 20%/Bear 8%",
    'tsla_share': "2025市占率9.5% (Bloomberg/InsideEVs)\n历史峰值17%(2021)→当前9.5%\nBase终年7%: 参考Intel CPU从90%→70%花8年\nBull 10%: FSD+Optimus生态锁定\nBear 5%: 比亚迪+新势力持续蚕食",
    'asp': "2025 ASP ~$38K (交付量173万辆 → 汽车收入$69.5B)\nBase终年$32K: Model 2拉低均价但Cybertruck/FSD提价\n参考: 丰田全球ASP ~$28K, 宝马 ~$55K",
    'storage_tam': "2025全球储能 ~$50B (BloombergNEF)\nAI数据中心需求+可再生能源并网=结构性驱动\nBase 30%: BNEF预测储能市场年增25-35%",
    'storage_share': "2025 ~25% (Tesla Megapack/Powerwall)\n上海Megapack工厂2025Q1投产, 产能翻倍\n竞争对手: 宁德时代、比亚迪、Fluence",
    'services': "2025A $12.5B, 随保有量线性增长\n超充网络对外开放+二手车+维修\nBase Y1-3 12%: 车队规模+非Tesla充电收入",
    'optimus': "2025试产~5000台, 2026目标5-10万台\nBase: 2027出货1万台, CAGR 60%(→2030)→40%(→2035)\nDeutsche Bank 2035E: 20万台×$50K=$10B (乐观)\n参考: 工业机器人市场CAGR ~15%",
    'gm': "历史: 25.6%(2022)→18.0%(2025), 降价压毛利\nBase 18%→21%: 能源占比提升(30%GM)+Optimus放量\n压制: 汽车价格战、Cybertruck爬坡成本\nBull 26%: 能源+Optimus占比超40%",
    'exp': "历史费用率: R&D 4-7% + SGA 5-6%\nBase 15%→13%: 收入增长快于费用\nBear 18%: AI/FSD/Optimus R&D刚性支出不降",
    'tax': "Base 18%: 美国企业税率21%+海外低税率+R&D抵免\n历史有效税率波动大(2023年负值-一次性项目)",
    'pe': "当前PE-TTM 405x (盈利低谷: EPS $1.08 vs 高峰$4.30)\n正常化EPS ~$5 → 正常化PE ~87x\nBase起始80x(正常化), 终年25x(增速~6%时合理)\n参考: Toyota 10x, Apple 30x, NVDA(高增长期)50x+",
    'shares': "3756M = yfinance sharesOutstanding\n假设不变(保守: Tesla有回购但亦有股权激励稀释)",
}

def _cagr_for(t, c1, c2, c3):
    return c1 if t <= 3 else (c2 if t <= 6 else c3)

def project_all():
    n = 10
    scenarios = {}
    for name in ['bear', 'base', 'bull']:
        p = ASSUMPTIONS[name]
        a, e, s, o = p['auto'], p['energy'], p['services'], p['optimus']
        last_rev = HIST_DATA['2025A']['revenue']
        rows = {t: {} for t in range(1, n+1)}

        prev_ev_tam = None; prev_storage_tam = None
        prev_svc_rev = None; prev_opt_units = None

        for t in range(1, n+1):
            # --- Automotive ---
            cagr_ev = _cagr_for(t, a['tam_cagr_1_3'], a['tam_cagr_4_6'], a['tam_cagr_7_10'])
            ev_tam = (a['ev_tam_2025'] * (1+cagr_ev)) if t==1 else prev_ev_tam*(1+cagr_ev)
            share = a['share_2025'] + (a['share_2035']-a['share_2025']) * t / n
            asp = a['asp_2025'] + (a['asp_2035']-a['asp_2025']) * t / n
            auto_rev = ev_tam * share * asp  # M units × $K → $B (10^6 × 10^3 = 10^9)
            prev_ev_tam = ev_tam

            # --- Energy ---
            cagr_st = _cagr_for(t, e['tam_cagr_1_3'], e['tam_cagr_4_6'], e['tam_cagr_7_10'])
            st_tam = (e['tam_2025']*(1+cagr_st)) if t==1 else prev_storage_tam*(1+cagr_st)
            st_share = e['share_2025'] + (e['share_2035']-e['share_2025'])*t/n
            energy_rev = st_tam * st_share
            prev_storage_tam = st_tam

            # --- Services ---
            cagr_sv = _cagr_for(t, s['cagr_1_3'], s['cagr_4_6'], s['cagr_7_10'])
            svc_rev = (s['rev_2025']*(1+cagr_sv)) if t==1 else prev_svc_rev*(1+cagr_sv)
            prev_svc_rev = svc_rev

            # --- Optimus (kicks in 2027, t=2 onwards) ---
            if t <= 1:
                opt_units = 0; opt_rev = 0; prev_opt_units = 0
            else:
                if t == 2:  # 2027
                    opt_units = o['units_2027']
                else:
                    cagr_o = o['unit_cagr_27_30'] if t <= 5 else o['unit_cagr_31_35']
                    opt_units = prev_opt_units * (1+cagr_o)
                opt_asp = o['asp_2027'] + (o['asp_2035']-o['asp_2027']) * max(0, t-2) / (n-2)
                opt_rev = opt_units * opt_asp  # M units × $K → $B (10^6 × 10^3 = 10^9)
                prev_opt_units = opt_units

            total_rev = auto_rev + energy_rev + svc_rev + opt_rev

            # Growth
            if t == 1:
                rev_growth = (total_rev - last_rev) / last_rev
            else:
                rev_growth = (total_rev - rows[t-1]['total_rev']) / rows[t-1]['total_rev']

            # Margins
            gm = p['gm_start'] + (p['gm_end']-p['gm_start'])*(t-1)/(n-1)
            gross_profit = total_rev * gm
            exp_r = p['exp_start'] + (p['exp_end']-p['exp_start'])*(t-1)/(n-1)
            opex = total_rev * exp_r
            ebit = gross_profit - opex
            ni = ebit * (1-p['tax_rate'])
            pe = p['pe_start'] + (p['pe_end']-p['pe_start'])*(t-1)/(n-1)
            mkt_cap = ni * pe
            price = mkt_cap * 1000 / p['shares_m']

            rows[t] = {
                'ev_tam': ev_tam, 'ev_share': share, 'asp': asp,
                'auto_rev': auto_rev,
                'storage_tam': st_tam, 'storage_share': st_share,
                'energy_rev': energy_rev,
                'svc_rev': svc_rev,
                'opt_units': opt_units, 'opt_asp': opt_asp if t>=2 else 0, 'opt_rev': opt_rev,
                'total_rev': total_rev, 'rev_growth': rev_growth,
                'gm': gm, 'gross_profit': gross_profit,
                'exp_ratio': exp_r, 'opex': opex,
                'ebit': ebit, 'tax_rate': p['tax_rate'],'net_income': ni,
                'ebit_margin': ebit/total_rev if total_rev else 0,
                'net_margin': ni/total_rev if total_rev else 0,
                'pe': pe, 'mkt_cap': mkt_cap,
                'shares_m': p['shares_m'], 'price': price,
            }

        terminal = rows[n]
        cagr_10y = (terminal['total_rev']/last_rev)**(1/n)-1
        if terminal['price'] > 0:
            annual_ret = (terminal['price']/VAL_DATA['current_price'])**(1/n)-1
        else:
            annual_ret = -1.0
        scenarios[name] = {
            'rows': rows,
            'terminal': {
                'year': 2035,
                'rev': terminal['total_rev'], 'gm': terminal['gm'],
                'ebit_margin': terminal['ebit_margin'],
                'net_income': terminal['net_income'],
                'mkt_cap': terminal['mkt_cap'], 'price': terminal['price'],
                'cagr': cagr_10y, 'annual_return': annual_ret, 'pe': terminal['pe'],
            }
        }
    return scenarios

def run_sanity(scenarios):
    checks = []
    n=10
    for s_name, s in scenarios.items():
        rows = s['rows']; name = s_name.upper()
        # Growth decay: allow up to 2 accelerations (CAGR block boundaries)
        g = [rows[t]['rev_growth'] for t in range(1,n+1)]
        accel_streak=0; max_accel_streak=0
        for i in range(1,len(g)):
            if g[i] > g[i-1]:
                accel_streak += 1; max_accel_streak = max(max_accel_streak, accel_streak)
            else:
                accel_streak = 0
        checks.append({'name': f'[{name}] 收入增速衰减', 'status': '✓' if max_accel_streak<3 else '⚠'})
        # GM range
        gms=[rows[t]['gm'] for t in range(1,n+1)]
        checks.append({'name': f'[{name}] 毛利率范围', 'status': '✓' if all(0.08<g<0.35 for g in gms) else '✗'})
        # EBIT margin
        ebm=[rows[t]['ebit_margin'] for t in range(1,n+1)]
        checks.append({'name': f'[{name}] EBIT利润率合理', 'status': '✓' if all(0<e<0.30 for e in ebm) else '✗'})
        # Revenue size check
        checks.append({'name': f'[{name}] 终年收入 vs 当前', 'bear': '', 'base': '', 'bull': '',
                       'status': f"${s['terminal']['rev']:.0f}B (CAGR {s['terminal']['cagr']:.1%})"})

    bt, bet, but = [scenarios[x]['terminal'] for x in ['bear','base','bull']]
    order_ok = but['rev']>bet['rev']>bt['rev']
    ni_ok = but['net_income']>bet['net_income']>bt['net_income']
    gap_bb = (bet['price']-bt['price'])/bt['price'] if bt['price'] > 0 else 999
    gap_bu = (but['price']-bet['price'])/bet['price'] if bet['price'] > 0 else 999
    checks.append({'name':'收入排序 Bull>Base>Bear',
                   'bear':f"${bt['rev']:.0f}B",'base':f"${bet['rev']:.0f}B",'bull':f"${but['rev']:.0f}B",
                   'status':'✓' if order_ok else '✗'})
    checks.append({'name':'净利排序 Bull>Base>Bear',
                   'bear':f"${bt['net_income']:.0f}B",'base':f"${bet['net_income']:.0f}B",'bull':f"${but['net_income']:.0f}B",
                   'status':'✓' if ni_ok else '✗'})
    checks.append({'name':f'情景差距 (Bear-Base {gap_bb:.0%}, Base-Bull {gap_bu:.0%})',
                   'status':'✓' if 0.2<gap_bb<2.5 and 0.2<gap_bu<2.5 else '⚠'})
    # External
    checks.append({'name':f'Base PE终年 {bet["pe"]:.0f}x vs Toyota 10x / Apple 30x',
                   'status':'✓ 合理区间' if 15<=bet['pe']<=35 else '⚠'})
    checks.append({'name':f'Base CAGR {bet["cagr"]:.1%} vs EV行业~12%',
                   'status':'✓' if 0.05<bet['cagr']<0.20 else '⚠'})
    return checks

# ===== Excel Builder =====
def style_header_row(ws, row, start, end, fill=HEADER_FILL, font=WHITE_BOLD):
    for c in range(start, end+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def style_data_cell(ws, row, col, font=BLACK_FONT, fill=None, num_fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font; cell.border = THIN_BORDER; cell.alignment = RIGHT_ALIGN
    if fill: cell.fill = fill
    if num_fmt: cell.number_format = num_fmt

def build_history(ws):
    ws.title = LABELS['sheet_hist']
    years = ['2021A','2022A','2023A','2024A','2025A']
    ws.cell(row=1,column=1,value='指标'); style_header_row(ws,1,1,len(years)+1)
    for i,y in enumerate(years):
        ws.cell(row=1,column=i+2,value=y)
        style_header_row(ws,1,i+2,i+2,fill=COL_HEADER_FILL,font=BLACK_BOLD)

    rows_def = [
        ('收入 ($B)', 'revenue', BLUE_FONT, INPUT_FILL, '#,##0.0'),
        ('收入增速', 'revenue_growth', BLACK_FONT, None, '0.0%'),
        ('毛利 ($B)', 'gross_profit', BLACK_FONT, None, '#,##0.0'),
        ('毛利率', 'gross_margin', BLACK_FONT, None, '0.0%'),
        ('研发费用 ($B)', 'rd', BLUE_FONT, INPUT_FILL, '#,##0.0'),
        ('研发费用率', 'rd_ratio', BLACK_FONT, None, '0.0%'),
        ('销售管理费用 ($B)', 'sga', BLUE_FONT, INPUT_FILL, '#,##0.0'),
        ('SGA费用率', 'sga_ratio', BLACK_FONT, None, '0.0%'),
        ('EBIT ($B)', 'ebit', BLACK_FONT, None, '#,##0.0'),
        ('EBIT利润率', 'ebit_margin', BLACK_FONT, None, '0.0%'),
        ('净利润 ($B)', 'net_income', BLUE_FONT, INPUT_FILL, '#,##0.0'),
        ('净利率', 'net_margin', BLACK_FONT, None, '0.0%'),
    ]
    for r,(label,key,font,fill,fmt) in enumerate(rows_def,2):
        ws.cell(row=r,column=1,value=label); ws.cell(row=r,column=1).font=BLACK_BOLD
        ws.cell(row=r,column=1).border=THIN_BORDER
        for c,y in enumerate(years,2):
            val=HIST_DATA[y].get(key)
            if val is not None: ws.cell(row=r,column=c,value=val)
            style_data_cell(ws,r,c,font=font,fill=fill,num_fmt=fmt)
    ws.column_dimensions['A'].width=26
    for c in range(2,len(years)+2): ws.column_dimensions[get_column_letter(c)].width=13

def build_segments_sheet(ws):
    ws.title = LABELS['sheet_segments']
    years = ['2021A','2022A','2023A','2024A','2025A']
    ws.cell(row=1,column=1,value='分部'); style_header_row(ws,1,1,len(years)+1)
    for i,y in enumerate(years):
        ws.cell(row=1,column=i+2,value=y); style_header_row(ws,1,i+2,i+2,fill=COL_HEADER_FILL,font=BLACK_BOLD)

    segs = [('汽车 ($B)','auto_revenue'),('能源存储 ($B)','energy_revenue'),
            ('服务及其他 ($B)','service_revenue'),('总收入 ($B)','revenue')]
    for r,(label,key) in enumerate(segs,2):
        ws.cell(row=r,column=1,value=label); ws.cell(row=r,column=1).font=BLACK_BOLD
        ws.cell(row=r,column=1).border=THIN_BORDER
        for c,y in enumerate(years,2):
            val=HIST_DATA[y].get(key)
            if val is not None: ws.cell(row=r,column=c,value=val)
            style_data_cell(ws,r,c,font=BLUE_FONT,fill=INPUT_FILL,num_fmt='#,##0.0')
    # Share rows
    for r_offset,(seg_name,seg_key) in enumerate([('汽车','auto_revenue'),('能源','energy_revenue'),('服务','service_revenue')]):
        r = r_offset+6
        ws.cell(row=r,column=1,value=f'{seg_name} 收入占比'); ws.cell(row=r,column=1).font=BLACK_BOLD
        ws.cell(row=r,column=1).border=THIN_BORDER
        for c,y in enumerate(years,2):
            rev=HIST_DATA[y].get(seg_key,0) or 0; total=HIST_DATA[y].get('revenue',1) or 1
            ws.cell(row=r,column=c,value=rev/total if total else 0)
            style_data_cell(ws,r,c,num_fmt='0.0%')
    ws.column_dimensions['A'].width=26
    for c in range(2,len(years)+2): ws.column_dimensions[get_column_letter(c)].width=13

def build_valuation_sheet(ws):
    ws.title = LABELS['sheet_val']
    ws.cell(row=1,column=1,value='指标'); style_header_row(ws,1,1,1)
    rows=[
        ('当前市值 ($B)', VAL_DATA['current_market_cap']/1e9, '#,##0.0'),
        ('当前股价 ($)', VAL_DATA['current_price'], '#,##0.00'),
        ('PE (TTM)', VAL_DATA['current_pe'], '0.0'),
        ('PE (Forward)', VAL_DATA['forward_pe'], '0.0'),
        ('流通股数 (M)', VAL_DATA['shares_outstanding']/1e6, '#,##0'),
        ('52周高 ($)', VAL_DATA['fifty_two_week_high'], '#,##0.00'),
        ('52周低 ($)', VAL_DATA['fifty_two_week_low'], '#,##0.00'),
        ('5年PE低', VAL_DATA['pe_5y_low'], '0.0'),
        ('5年PE中位', VAL_DATA['pe_5y_median'], '0.0'),
        ('5年PE高', VAL_DATA['pe_5y_high'], '0.0'),
    ]
    for r,(label,val,fmt) in enumerate(rows,2):
        ws.cell(row=r,column=1,value=label); ws.cell(row=r,column=1).font=BLACK_BOLD
        ws.cell(row=r,column=1).border=THIN_BORDER
        ws.cell(row=r,column=2,value=val); style_data_cell(ws,r,2,font=BLUE_FONT,fill=INPUT_FILL,num_fmt=fmt)
    ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=14

def build_assumptions(ws):
    ws.title = LABELS['sheet_assumptions']
    # B1: scenario toggle
    ws.cell(row=1,column=1,value=LABELS['active_scenario'])
    ws.cell(row=1,column=1).font=WHITE_BOLD; ws.cell(row=1,column=1).fill=HEADER_FILL
    ws.cell(row=1,column=1).border=THIN_BORDER
    ws.cell(row=1,column=2,value=2)
    ws.cell(row=1,column=2).font=BLACK_BOLD; ws.cell(row=1,column=2).fill=TOGGLE_FILL
    ws.cell(row=1,column=2).border=THIN_BORDER
    dv=DataValidation(type="list",formula1='"1,2,3"',allow_blank=False)
    dv.error="1=Bear, 2=Base, 3=Bull"; dv.prompt="1=Bear, 2=Base, 3=Bull"
    ws.add_data_validation(dv); dv.add(ws["B1"])

    # Headers
    ws.cell(row=2,column=1,value='参数'); ws.cell(row=2,column=2,value='Bear')
    ws.cell(row=2,column=3,value='Base'); ws.cell(row=2,column=4,value='Bull')
    ws.cell(row=2,column=5,value='活跃值'); ws.cell(row=2,column=6,value=LABELS['source_label'])
    style_header_row(ws,2,1,6)
    style_header_row(ws,2,2,4,fill=COL_HEADER_FILL,font=BLACK_BOLD)
    style_header_row(ws,2,5,5,fill=OUTPUT_FILL,font=BLACK_BOLD)

    # === Assumption rows ===
    # (label, key_path, fmt, source_key)
    # key_path: "segment.param" or "param"
    assumption_rows = [
        (f"═══ {LABELS['auto']} ═══", None, None, None),
        (LABELS['ev_tam_2025'], 'auto.ev_tam_2025', '#,##0.0', 'ev_tam_2025'),
        (LABELS['ev_tam_cagr_13'], 'auto.tam_cagr_1_3', '0.0%', 'ev_tam_cagr'),
        (LABELS['ev_tam_cagr_46'], 'auto.tam_cagr_4_6', '0.0%', 'ev_tam_cagr'),
        (LABELS['ev_tam_cagr_710'], 'auto.tam_cagr_7_10', '0.0%', 'ev_tam_cagr'),
        (LABELS['tsla_share_2025'], 'auto.share_2025', '0.0%', 'tsla_share'),
        (LABELS['tsla_share_2035'], 'auto.share_2035', '0.0%', 'tsla_share'),
        (LABELS['asp_2025'], 'auto.asp_2025', '#,##0.0', 'asp'),
        (LABELS['asp_2035'], 'auto.asp_2035', '#,##0.0', 'asp'),
        (f"═══ {LABELS['energy']} ═══", None, None, None),
        (LABELS['storage_tam_2025'], 'energy.tam_2025', '#,##0.0', 'storage_tam'),
        (LABELS['storage_cagr_13'], 'energy.tam_cagr_1_3', '0.0%', 'storage_tam'),
        (LABELS['storage_cagr_46'], 'energy.tam_cagr_4_6', '0.0%', 'storage_tam'),
        (LABELS['storage_cagr_710'], 'energy.tam_cagr_7_10', '0.0%', 'storage_tam'),
        (LABELS['storage_share_2025'], 'energy.share_2025', '0.0%', 'storage_share'),
        (LABELS['storage_share_2035'], 'energy.share_2035', '0.0%', 'storage_share'),
        (f"═══ {LABELS['services']} ═══", None, None, None),
        (LABELS['svc_rev_2025'], 'services.rev_2025', '#,##0.0', 'services'),
        (LABELS['svc_cagr_13'], 'services.cagr_1_3', '0.0%', 'services'),
        (LABELS['svc_cagr_46'], 'services.cagr_4_6', '0.0%', 'services'),
        (LABELS['svc_cagr_710'], 'services.cagr_7_10', '0.0%', 'services'),
        (f"═══ {LABELS['optimus']} ═══", None, None, None),
        (LABELS['opt_units_2027'], 'optimus.units_2027', '#,##0.000', 'optimus'),
        (LABELS['opt_unit_cagr_2730'], 'optimus.unit_cagr_27_30', '0.0%', 'optimus'),
        (LABELS['opt_unit_cagr_3135'], 'optimus.unit_cagr_31_35', '0.0%', 'optimus'),
        (LABELS['opt_asp_2027'], 'optimus.asp_2027', '#,##0.0', 'optimus'),
        (LABELS['opt_asp_2035'], 'optimus.asp_2035', '#,##0.0', 'optimus'),
        (f"═══ 利润率 ═══", None, None, None),
        (LABELS['gm_start'], 'gm_start', '0.0%', 'gm'),
        (LABELS['gm_end'], 'gm_end', '0.0%', 'gm'),
        (LABELS['exp_start'], 'exp_start', '0.0%', 'exp'),
        (LABELS['exp_end'], 'exp_end', '0.0%', 'exp'),
        (LABELS['tax_rate'], 'tax_rate', '0.0%', 'tax'),
        (f"═══ 估值 ═══", None, None, None),
        (LABELS['pe_start'], 'pe_start', '0.0', 'pe'),
        (LABELS['pe_end'], 'pe_end', '0.0', 'pe'),
        (LABELS['shares_m'], 'shares_m', '#,##0', 'shares'),
    ]

    row_map = {}
    r=3
    for label, key_path, fmt, source_key in assumption_rows:
        if key_path is None:
            for c in range(1,7):
                ws.cell(row=r,column=c).fill=SEGMENT_FILL
                ws.cell(row=r,column=c).border=THIN_BORDER
            ws.cell(row=r,column=1,value=label); ws.cell(row=r,column=1).font=BLACK_BOLD
            r+=1; continue

        row_map[key_path]=r
        ws.cell(row=r,column=1,value=label); ws.cell(row=r,column=1).font=BLACK_BOLD
        ws.cell(row=r,column=1).border=THIN_BORDER

        parts=key_path.split('.')
        for ci,sc in enumerate(['bear','base','bull'],2):
            if len(parts)==2:
                val=ASSUMPTIONS[sc].get(parts[0],{}).get(parts[1])
            else:
                val=ASSUMPTIONS[sc].get(parts[0])
            if val is not None: ws.cell(row=r,column=ci,value=val)
            style_data_cell(ws,r,ci,font=BLUE_FONT,fill=INPUT_FILL,num_fmt=fmt)

        ws.cell(row=r,column=5,value=f'=INDEX(B{r}:D{r},$B$1)')
        style_data_cell(ws,r,5,font=GREEN_FONT,fill=OUTPUT_FILL,num_fmt=fmt)

        src=SOURCES.get(source_key,'') if source_key else ''
        ws.cell(row=r,column=6,value=src)
        ws.cell(row=r,column=6).font=GRAY_FONT; ws.cell(row=r,column=6).fill=NOTE_FILL
        ws.cell(row=r,column=6).border=THIN_BORDER; ws.cell(row=r,column=6).alignment=WRAP_ALIGN
        r+=1

    ws.column_dimensions['A'].width=32; ws.column_dimensions['F'].width=55
    for c in 'BCDE': ws.column_dimensions[c].width=12
    return row_map

def build_forecast(ws, row_map):
    ws.title = LABELS['sheet_forecast']
    proj_years=[f'{y}E' for y in range(2026,2036)]
    n=len(proj_years)
    CURR_PRICE=VAL_DATA['current_price']

    ws.cell(row=1,column=1,value='指标')
    for i,y in enumerate(proj_years):
        ws.cell(row=1,column=i+2,value=y)
    style_header_row(ws,1,1,n+1)
    style_header_row(ws,1,2,n+1,fill=COL_HEADER_FILL,font=BLACK_BOLD)

    def A(key_path):
        if key_path in row_map:
            return f"'{LABELS['sheet_assumptions']}'!$E${row_map[key_path]}"
        return "0"

    # Abbreviated refs
    EV_TAM=A('auto.ev_tam_2025'); EV_C13=A('auto.tam_cagr_1_3'); EV_C46=A('auto.tam_cagr_4_6'); EV_C710=A('auto.tam_cagr_7_10')
    EV_SH_ST=A('auto.share_2025'); EV_SH_END=A('auto.share_2035'); ASP_ST=A('auto.asp_2025'); ASP_END=A('auto.asp_2035')
    ST_TAM=A('energy.tam_2025'); ST_C13=A('energy.tam_cagr_1_3'); ST_C46=A('energy.tam_cagr_4_6'); ST_C710=A('energy.tam_cagr_7_10')
    ST_SH_ST=A('energy.share_2025'); ST_SH_END=A('energy.share_2035')
    SVC_REV=A('services.rev_2025'); SVC_C13=A('services.cagr_1_3'); SVC_C46=A('services.cagr_4_6'); SVC_C710=A('services.cagr_7_10')
    OPT_U27=A('optimus.units_2027'); OPT_UC2730=A('optimus.unit_cagr_27_30'); OPT_UC3135=A('optimus.unit_cagr_31_35')
    OPT_A27=A('optimus.asp_2027'); OPT_A35=A('optimus.asp_2035')
    GM_ST=A('gm_start'); GM_END=A('gm_end'); EXP_ST=A('exp_start'); EXP_END=A('exp_end')
    TAX=A('tax_rate'); PE_ST=A('pe_start'); PE_END=A('pe_end'); SH=A('shares_m')

    # === Row 2: Automotive section header ===
    r=2; ws.cell(row=r,column=1,value=f"═══ {LABELS['auto']} ═══")
    ws.cell(row=r,column=1).font=WHITE_BOLD; ws.cell(row=r,column=1).fill=HEADER_FILL
    for c in range(2,n+2): ws.cell(row=r,column=c).fill=HEADER_FILL

    # R3: EV TAM
    r=3; ws.cell(row=r,column=1,value='全球EV销量 (M辆)'); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        cagr_f=EV_C13 if i<3 else (EV_C46 if i<6 else EV_C710)
        if i==0: ws.cell(row=r,column=col,value=f"={EV_TAM}*(1+{cagr_f})")
        else: ws.cell(row=r,column=col,value=f"={get_column_letter(col-1)}{r}*(1+{cagr_f})")
        style_data_cell(ws,r,col,num_fmt='#,##0.0')

    # R4: Tesla EV share
    r=4; ws.cell(row=r,column=1,value='Tesla EV市占率'); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; ws.cell(row=r,column=col,value=f"={EV_SH_ST}+({EV_SH_END}-{EV_SH_ST})*{i+1}/{n}")
        style_data_cell(ws,r,col,num_fmt='0.0%')

    # R5: ASP
    r=5; ws.cell(row=r,column=1,value='ASP ($K)'); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; ws.cell(row=r,column=col,value=f"={ASP_ST}+({ASP_END}-{ASP_ST})*{i+1}/{n}")
        style_data_cell(ws,r,col,num_fmt='#,##0.0')

    # R6: Auto revenue
    r=6; ws.cell(row=r,column=1,value=f"{LABELS['auto']} 收入 ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}3*{c}4*{c}5")
        style_data_cell(ws,r,col,fill=OUTPUT_FILL,num_fmt='#,##0.0')

    # === Energy ===
    r=7; ws.cell(row=r,column=1,value=f"═══ {LABELS['energy']} ═══")
    ws.cell(row=r,column=1).font=WHITE_BOLD; ws.cell(row=r,column=1).fill=HEADER_FILL
    for c in range(2,n+2): ws.cell(row=r,column=c).fill=HEADER_FILL

    r=8; ws.cell(row=r,column=1,value='全球储能TAM ($B)'); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        cagr_f=ST_C13 if i<3 else (ST_C46 if i<6 else ST_C710)
        if i==0: ws.cell(row=r,column=col,value=f"={ST_TAM}*(1+{cagr_f})")
        else: ws.cell(row=r,column=col,value=f"={get_column_letter(col-1)}{r}*(1+{cagr_f})")
        style_data_cell(ws,r,col,num_fmt='#,##0.0')

    r=9; ws.cell(row=r,column=1,value='Tesla储能市占率'); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; ws.cell(row=r,column=col,value=f"={ST_SH_ST}+({ST_SH_END}-{ST_SH_ST})*{i+1}/{n}")
        style_data_cell(ws,r,col,num_fmt='0.0%')

    r=10; ws.cell(row=r,column=1,value=f"{LABELS['energy']} 收入 ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}8*{c}9")
        style_data_cell(ws,r,col,fill=OUTPUT_FILL,num_fmt='#,##0.0')

    # === Services ===
    r=11; ws.cell(row=r,column=1,value=f"═══ {LABELS['services']} ═══")
    ws.cell(row=r,column=1).font=WHITE_BOLD; ws.cell(row=r,column=1).fill=HEADER_FILL
    for c in range(2,n+2): ws.cell(row=r,column=c).fill=HEADER_FILL

    r=12; ws.cell(row=r,column=1,value=f"{LABELS['services']} 收入 ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        cagr_f=SVC_C13 if i<3 else (SVC_C46 if i<6 else SVC_C710)
        if i==0: ws.cell(row=r,column=col,value=f"={SVC_REV}*(1+{cagr_f})")
        else: ws.cell(row=r,column=col,value=f"={get_column_letter(col-1)}{r}*(1+{cagr_f})")
        style_data_cell(ws,r,col,fill=OUTPUT_FILL,num_fmt='#,##0.0')

    # === Optimus ===
    r=13; ws.cell(row=r,column=1,value=f"═══ {LABELS['optimus']} ═══")
    ws.cell(row=r,column=1).font=WHITE_BOLD; ws.cell(row=r,column=1).fill=HEADER_FILL
    for c in range(2,n+2): ws.cell(row=r,column=c).fill=HEADER_FILL

    r=14; ws.cell(row=r,column=1,value='出货量 (M台)'); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        if i==0: # 2026: 0
            ws.cell(row=r,column=col,value=0)
        elif i==1: # 2027: first year
            ws.cell(row=r,column=col,value=f"={OPT_U27}")
        else:
            cagr_o=OPT_UC2730 if i+1<=5 else OPT_UC3135
            ws.cell(row=r,column=col,value=f"={get_column_letter(col-1)}{r}*(1+{cagr_o})")
        style_data_cell(ws,r,col,num_fmt='#,##0.000')

    r=15; ws.cell(row=r,column=1,value='ASP ($K)'); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2
        if i<1:
            ws.cell(row=r,column=col,value=0)
        elif i==1:
            ws.cell(row=r,column=col,value=f"={OPT_A27}")
        else:
            ws.cell(row=r,column=col,value=f"={OPT_A27}+({OPT_A35}-{OPT_A27})*{i-1}/{n-2}")
        style_data_cell(ws,r,col,num_fmt='#,##0.0')

    r=16; ws.cell(row=r,column=1,value=f"{LABELS['optimus']} 收入 ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}14*{c}15")
        style_data_cell(ws,r,col,fill=OUTPUT_FILL,num_fmt='#,##0.0')

    # === Total ===
    r=17; ws.cell(row=r,column=1,value=f"═══ {LABELS['total']} ═══")
    ws.cell(row=r,column=1).font=WHITE_BOLD; ws.cell(row=r,column=1).fill=HEADER_FILL
    for c in range(2,n+2): ws.cell(row=r,column=c).fill=HEADER_FILL

    r=18; ws.cell(row=r,column=1,value=f"{LABELS['total']} ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}6+{c}10+{c}12+{c}16")
        style_data_cell(ws,r,col,font=BLACK_BOLD,fill=OUTPUT_FILL,num_fmt='#,##0.0')

    r=19; ws.cell(row=r,column=1,value=LABELS['fc_rev_growth']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        if i==0:
            last_rev=HIST_DATA['2025A']['revenue']
            ws.cell(row=r,column=col,value=f"=({c}18-{last_rev})/{last_rev}")
        else:
            pc=get_column_letter(col-1)
            ws.cell(row=r,column=col,value=f"=({c}18-{pc}18)/{pc}18")
        style_data_cell(ws,r,col,num_fmt='0.0%')

    # === Profit build-up ===
    r=21; ws.cell(row=r,column=1,value="═══ 利润推导 ═══")
    ws.cell(row=r,column=1).font=WHITE_BOLD; ws.cell(row=r,column=1).fill=HEADER_FILL
    for c in range(2,n+2): ws.cell(row=r,column=c).fill=HEADER_FILL

    r=22; ws.cell(row=r,column=1,value=LABELS['fc_gm']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2
        ws.cell(row=r,column=col,value=f"={GM_ST}+({GM_END}-{GM_ST})*{i}/{n-1}")
        style_data_cell(ws,r,col,num_fmt='0.0%')

    r=23; ws.cell(row=r,column=1,value=f"{LABELS['fc_gross_profit']} ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}18*{c}22")
        style_data_cell(ws,r,col,num_fmt='#,##0.0')

    r=24; ws.cell(row=r,column=1,value=LABELS['fc_exp_ratio']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2
        ws.cell(row=r,column=col,value=f"={EXP_ST}+({EXP_END}-{EXP_ST})*{i}/{n-1}")
        style_data_cell(ws,r,col,num_fmt='0.0%')

    r=25; ws.cell(row=r,column=1,value=f"{LABELS['fc_opex']} ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}18*{c}24")
        style_data_cell(ws,r,col,num_fmt='#,##0.0')

    r=26; ws.cell(row=r,column=1,value=LABELS['fc_ebit']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}23-{c}25")
        style_data_cell(ws,r,col,num_fmt='#,##0.0')

    r=27; ws.cell(row=r,column=1,value=LABELS['fc_tax_rate']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2
        ws.cell(row=r,column=col,value=f"={TAX}")
        style_data_cell(ws,r,col,font=GREEN_FONT,num_fmt='0.0%')

    r=28; ws.cell(row=r,column=1,value=f"{LABELS['fc_net_income']} ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}26*(1-{c}27)")
        style_data_cell(ws,r,col,fill=OUTPUT_FILL,num_fmt='#,##0.0')

    # === Valuation ===
    r=30; ws.cell(row=r,column=1,value="═══ 估值 ═══")
    ws.cell(row=r,column=1).font=WHITE_BOLD; ws.cell(row=r,column=1).fill=HEADER_FILL
    for c in range(2,n+2): ws.cell(row=r,column=c).fill=HEADER_FILL

    r=31; ws.cell(row=r,column=1,value=LABELS['fc_target_pe']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2
        ws.cell(row=r,column=col,value=f"={PE_ST}+({PE_END}-{PE_ST})*{i}/{n-1}")
        style_data_cell(ws,r,col,num_fmt='0.0')

    r=32; ws.cell(row=r,column=1,value=f"{LABELS['fc_implied_mkt_cap']} ($B)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}28*{c}31")
        style_data_cell(ws,r,col,fill=OUTPUT_FILL,num_fmt='#,##0.0')

    r=33; ws.cell(row=r,column=1,value=LABELS['fc_shares']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2
        ws.cell(row=r,column=col,value=f"={SH}")
        style_data_cell(ws,r,col,font=GREEN_FONT,num_fmt='#,##0')

    r=34; ws.cell(row=r,column=1,value=f"{LABELS['fc_implied_price']} ($)"); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"={c}32*1000/{c}33")
        style_data_cell(ws,r,col,font=BLACK_BOLD,fill=OUTPUT_FILL,num_fmt='#,##0.00')

    r=35; ws.cell(row=r,column=1,value=LABELS['fc_premium']); ws.cell(row=r,column=1).font=BLACK_BOLD; ws.cell(row=r,column=1).border=THIN_BORDER
    for i in range(n):
        col=i+2; c=get_column_letter(col)
        ws.cell(row=r,column=col,value=f"=({c}34-{CURR_PRICE})/{CURR_PRICE}")
        style_data_cell(ws,r,col,num_fmt='0.0%')

    ws.column_dimensions['A'].width=32
    for c in range(2,n+2): ws.column_dimensions[get_column_letter(c)].width=13

def build_sanity_sheet(ws, checks):
    ws.title = LABELS['sheet_sanity']
    ws.cell(row=1,column=1,value='检验项'); ws.cell(row=1,column=2,value='Bear')
    ws.cell(row=1,column=3,value='Base'); ws.cell(row=1,column=4,value='Bull')
    ws.cell(row=1,column=5,value='结果')
    style_header_row(ws,1,1,5)
    for ri,ck in enumerate(checks,2):
        ws.cell(row=ri,column=1,value=ck['name']); ws.cell(row=ri,column=1).font=BLACK_BOLD; ws.cell(row=ri,column=1).border=THIN_BORDER
        for ci,k in enumerate(['bear','base','bull'],2):
            val=ck.get(k,'')
            if val: ws.cell(row=ri,column=ci,value=val)
            style_data_cell(ws,ri,ci)
        ws.cell(row=ri,column=5,value=ck['status'])
        status_font=BLACK_FONT
        if ck['status'].startswith('✗'): status_font=Font(color="FF0000",bold=True)
        elif ck['status'].startswith('⚠'): status_font=Font(color="FF8C00",bold=True)
        style_data_cell(ws,ri,5,font=status_font)
    ws.column_dimensions['A'].width=40
    for c in 'BCDE': ws.column_dimensions[c].width=16

def build_fy_sheet(ws):
    ws.title = LABELS['sheet_fy']
    ws.cell(row=1,column=1,value='Tesla 财年信息').font=Font(size=14,bold=True,color='1F4E79')
    ws.merge_cells('A1:C1')
    info=[('FY结束月','12月'),('FY结束日','12月31日'),('与自然年一致','是'),
          ('映射规则','FY=自然年，无需转换'),('说明','Tesla财年结束于12月31日，与自然年完全一致。所有FY标签直接对应自然年。')]
    for i,(k,v) in enumerate(info,3):
        ws.cell(row=i,column=1,value=k).font=BLACK_BOLD
        ws.cell(row=i,column=2,value=v)
    ws.column_dimensions['A'].width=20; ws.column_dimensions['B'].width=60

def main():
    parser=argparse.ArgumentParser(description='Tesla Coverage Model Builder')
    parser.add_argument('--output-dir',default='./out',help='输出目录')
    args=parser.parse_args()

    scenarios=project_all()
    checks=run_sanity(scenarios)

    # Print summary
    for name in ['bear','base','bull']:
        t=scenarios[name]['terminal']
        print(f"\n{'='*50}")
        print(f"  {name.upper()} CASE — 终年 2035E")
        print(f"{'='*50}")
        print(f"  收入:        ${t['rev']:>10,.1f}B  (CAGR: {t['cagr']:.1%})")
        print(f"  毛利率:      {t['gm']:>10.1%}")
        print(f"  EBIT利润率:  {t['ebit_margin']:>10.1%}")
        print(f"  净利润:      ${t['net_income']:>10,.1f}B")
        print(f"  终年PE:      {t['pe']:>10.1f}x")
        print(f"  隐含市值:    ${t['mkt_cap']:>10,.1f}B")
        print(f"  隐含股价:    ${t['price']:>10,.1f}")
        print(f"  vs 当前:     {(t['price']/VAL_DATA['current_price']-1):>10.1%}")
        print(f"  年化回报:    {t['annual_return']:>10.1%}")

    # Base case annual details
    base_rows=scenarios['base']['rows']
    print(f"\n{'='*80}")
    print(f"  BASE CASE — 逐年分部明细")
    print(f"{'='*80}")
    print(f"  {'Year':>6} {'Auto($B)':>10} {'Energy($B)':>12} {'Svc($B)':>10} {'Optimus($B)':>12} {'Total($B)':>10} {'Gr%':>7} {'GM%':>6} {'NI($B)':>10} {'PE':>5} {'Price':>8}")
    print(f"  {'-'*78}")
    for t in range(1,11):
        r=base_rows[t]; yr=2025+t
        print(f"  {yr}E ${r['auto_rev']:>9,.1f}B ${r['energy_rev']:>11,.1f}B ${r['svc_rev']:>9,.1f}B ${r['opt_rev']:>11,.1f}B ${r['total_rev']:>9,.1f}B {r['rev_growth']*100:>6.1f}% {r['gm']*100:>5.1f}% ${r['net_income']:>9,.1f}B {r['pe']:>4.1f}x ${r['price']:>7,.1f}")

    # Print sanity
    print(f"\n{'='*60}")
    print(f"  合理性检验")
    print(f"{'='*60}")
    for ck in checks:
        vals=' | '.join(str(ck.get(k,'')) for k in ['bear','base','bull'] if ck.get(k))
        line=f"  {ck['name']:<45}"
        if vals: line+=f" {vals}"
        line+=f"  {ck['status']}"
        print(line)

    # Build Excel
    wb=Workbook()
    # Sheet 1: Musk Empire Overview
    ws_overview=wb.active; build_musk_overview(ws_overview)
    # Tesla sheets
    ws1=wb.create_sheet(); build_history(ws1)
    ws2=wb.create_sheet(); build_segments_sheet(ws2)
    ws3=wb.create_sheet(); build_valuation_sheet(ws3)
    ws_fy=wb.create_sheet(); build_fy_sheet(ws_fy)
    ws_asm=wb.create_sheet(); row_map=build_assumptions(ws_asm)
    ws_fc=wb.create_sheet(); build_forecast(ws_fc, row_map)
    ws_ck=wb.create_sheet(); build_sanity_sheet(ws_ck, checks)
    # SpaceX & Neuralink
    ws_spx=wb.create_sheet(); build_spacex_sheet(ws_spx)
    ws_nlk=wb.create_sheet(); build_neuralink_sheet(ws_nlk)

    output_dir=Path(args.output_dir); output_dir.mkdir(parents=True,exist_ok=True)
    out_path=output_dir/'Musk_Coverage_Model.xlsx'
    wb.save(str(out_path))
    print(f"\n✓ Excel 已生成: {out_path}")

# ===== Musk Empire Overview =====
def build_musk_overview(ws):
    ws.title = 'Musk帝国总览'
    TITLE_FONT = Font(size=16, bold=True, color='1F4E79')
    SUB_FONT = Font(size=12, bold=True, color='1F4E79')

    r=1; ws.cell(row=r,column=1,value='Musk 帝国 Coverage Model').font=TITLE_FONT
    ws.merge_cells('A1:G1')
    r=2; ws.cell(row=r,column=1,value='Elon Musk 旗下核心企业估值覆盖 | 2026年5月').font=Font(size=11,color='666666')
    ws.merge_cells('A2:G2')

    # === Summary Table ===
    r=4
    headers=['公司','代码/状态','估值 ($B)','2025收入 ($B)','收入增速','阶段','核心驱动']
    for c,h in enumerate(headers,1):
        ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,len(headers))

    companies=[
        ['Tesla','TSLA (公开)','$1,628','$94.8','-3%','成熟+转型','EV + 能源 + Optimus'],
        ['SpaceX','私有 (IPO预计2026)','~$800','$15.5-18.7','~+40%','高速增长','Starlink + 发射 + xAI'],
        ['  └ Starlink','SpaceX子业务','—','$11.4-12.3','~+50%','高速增长','卫星互联网 9M+用户'],
        ['  └ xAI','2026.2并入SpaceX','$230(收购前)','~$0.5','早期','早期爆发','Grok AI + Colossus超算'],
        ['Neuralink','私有','$9','$0 (临床期)','N/A','临床阶段','BCI 脑机接口'],
        ['Optimus','Tesla旗下','含在TSLA内','~$0 (试产)','N/A','试产阶段','人形机器人'],
    ]
    for i,row_data in enumerate(companies):
        r=5+i
        for c,val in enumerate(row_data,1):
            ws.cell(row=r,column=c,value=val)
            ws.cell(row=r,column=c).border=THIN_BORDER
            ws.cell(row=r,column=c).font=BLACK_FONT
            if c==1: ws.cell(row=r,column=c).font=BLACK_BOLD

    # === Key Takeaways ===
    r=12
    ws.cell(row=r,column=1,value='核心结论').font=SUB_FONT
    r=13
    takeaways=[
        '1. Tesla 当前估值 ($1.6T) 定价了最乐观情景——Base case 合理价仅 $95',
        '2. SpaceX 是 Musk 帝国中基本面最扎实的资产——Starlink 年运营利润 ~$44B，2026年IPO目标 $1.75-2T',
        '3. xAI 已被 SpaceX 收购，合并后形成「发射+互联网+AI」三位一体，但烧钱速度极快 (~$1B/月)',
        '4. Neuralink 仍处于临床早期 (12例植入)，2029年FDA审批是关键里程碑，当前$9B估值已是BCI赛道最高',
        '5. Optimus 是最大的不确定性——从$0到可能$10B+收入，但量产爬坡和技术成熟度仍存疑',
        '6. Musk 帝国的核心风险：跨公司资金依赖 (SpaceX→xAI)、个人关键人风险、监管不确定性',
    ]
    for t in takeaways:
        ws.cell(row=r,column=1,value=t).font=Font(size=10)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
        r+=1

    # === Musk Net Worth Tie-in ===
    r+=2
    ws.cell(row=r,column=1,value='Musk 身价与公司估值关联').font=SUB_FONT
    r+=1
    ws.cell(row=r,column=1,value='Musk持股: Tesla ~13%, SpaceX ~42%, xAI ~40% (合并前), Neuralink ~50%+ (估算)').font=Font(size=10,color='666666')
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
    r+=1
    ws.cell(row=r,column=1,value='身价估算: Tesla $212B + SpaceX $336B + 其他 ~$20B ≈ $560B+ (全球首富)').font=Font(size=10,color='666666')
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)

    ws.column_dimensions['A'].width=22
    for c in 'BCDEFG': ws.column_dimensions[c].width=18


# ===== SpaceX =====
def build_spacex_sheet(ws):
    ws.title = 'SpaceX'
    TITLE_FONT = Font(size=14, bold=True, color='1F4E79')
    r=1; ws.cell(row=r,column=1,value='SpaceX — 私有公司估值模型 (含xAI)').font=TITLE_FONT
    ws.merge_cells('A1:G1')

    # === Revenue History ===
    r=3; ws.cell(row=r,column=1,value='收入估算 ($B)').font=Font(size=12,bold=True)
    r=4
    yrs=['2020A','2021A','2022A','2023A','2024E','2025E']
    ws.cell(row=r,column=1,value='分部'); style_header_row(ws,r,1,len(yrs)+1)
    for i,y in enumerate(yrs):
        ws.cell(row=r,column=i+2,value=y); style_header_row(ws,r,i+2,i+2,fill=COL_HEADER_FILL,font=BLACK_BOLD)

    spx_data={
        'Starlink': [0,0.2,1.0,4.5,8.0,11.9],
        '发射服务': [1.8,2.5,3.6,4.2,4.2,4.5],
        '其他': [0.2,0.3,0.3,0.3,0.3,0.3],
        '总收入': [2.0,3.0,4.9,9.0,12.5,16.7],
    }
    ri=5
    for label,vals in spx_data.items():
        ws.cell(row=ri,column=1,value=label); ws.cell(row=ri,column=1).font=BLACK_BOLD; ws.cell(row=ri,column=1).border=THIN_BORDER
        for ci,v in enumerate(vals,2):
            ws.cell(row=ri,column=ci,value=v)
            style_data_cell(ws,ri,ci,font=BLUE_FONT,fill=INPUT_FILL,num_fmt='#,##0.0')
        ri+=1

    # === Starlink Metrics ===
    r=10; ws.cell(row=r,column=1,value='Starlink 关键指标').font=Font(size=12,bold=True)
    r=11
    headers=['指标','2022A','2023A','2024A','2025E','2026E','2030E (Base)']
    for c,h in enumerate(headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,len(headers))
    starlink_metrics=[
        ['活跃用户 (M)',0.5,2.3,4.6,9.0,15.0,50.0],
        ['覆盖国家',40,70,120,155,180,200],
        ['在轨卫星',3000,5000,7000,9000,12000,20000],
        ['ARPU ($/月)',100,100,105,110,110,120],
        ['估算收入 ($B)',0.6,2.8,5.8,11.9,19.8,72.0],
        ['运营利润率','负','~20%','~35%','~45%','~50%','~55%'],
    ]
    for ri,row_d in enumerate(starlink_metrics,12):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # === Valuation History ===
    r=19; ws.cell(row=r,column=1,value='估值里程碑').font=Font(size=12,bold=True)
    r=20
    v_headers=['时间','估值 ($B)','每股 ($)','事件']
    for c,h in enumerate(v_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,4)
    val_hist=[
        ['2023.12','$180','$85','Tender Offer'],
        ['2024.06','$210','$98','二级交易'],
        ['2024.12','$350-400','$212','Tender Offer (Starlink盈利确认)'],
        ['2025年中','$400','—','Bloomberg报道'],
        ['2025.12','$800','$421','二级交易 (xAI合并预期)'],
        ['2026 (IPO目标)','$1,750-2,000','—','Nasdaq上市，或成史上最大IPO'],
    ]
    for ri,row_d in enumerate(val_hist,21):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # === xAI Note ===
    r=28; ws.cell(row=r,column=1,value='xAI (已并入SpaceX)').font=Font(size=12,bold=True,color='FF0000')
    r=29
    ws.cell(row=r,column=1,value='2026年2月，SpaceX以$1.25T合并估值全资收购xAI。xAI此前在2025年1月以$230B估值完成$20B Series E。')
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
    r=31
    xai_headers=['指标','2024E','2025E','2026E','2029E (目标)']
    for c,h in enumerate(xai_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,5)
    xai_data=[
        ['Grok月活用户','—','64M','150M+','500M+'],
        ['GPU集群规模','—','200K+','500K+','1M+'],
        ['收入 ($B)','~0.1','~0.5','~2','~14'],
        ['估值 ($B)','—','80→230','并入SpaceX','—'],
    ]
    for ri,row_d in enumerate(xai_data,32):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # === Bear/Base/Bull for SpaceX ===
    r=37; ws.cell(row=r,column=1,value='SpaceX 三情景分析 (2030E)').font=Font(size=12,bold=True)
    r=38
    s_headers=['指标','Bear','Base','Bull','备注']
    for c,h in enumerate(s_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,5)
    spx_scenarios=[
        ['Starlink用户 2030E','25M','50M','100M','渗透率5-20%'],
        ['发射次数/年','100','200','300','Starship复用是关键'],
        ['总收入 2030E ($B)','$45','$90','$180',''],
        ['IPO估值 ($T)','$0.8','$1.8','$3.5','2026年IPO为基准'],
        ['核心风险','Starship失败\n竞争加剧','星舰进度慢\nDTC不及预期','全面兑现\n垄断卫星互联网'],
    ]
    for ri,row_d in enumerate(spx_scenarios,39):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            ws.cell(row=ri,column=ci).alignment=WRAP_ALIGN
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    ws.column_dimensions['A'].width=22
    for c in 'BCDEFG': ws.column_dimensions[c].width=16


# ===== Neuralink =====
def build_neuralink_sheet(ws):
    ws.title = 'Neuralink'
    TITLE_FONT = Font(size=14, bold=True, color='1F4E79')
    r=1; ws.cell(row=r,column=1,value='Neuralink — 临床阶段BCI公司').font=TITLE_FONT
    ws.merge_cells('A1:G1')
    r=2; ws.cell(row=r,column=1,value='当前无收入，按里程碑+市场TAM折现估值').font=Font(size=10,color='666666')
    ws.merge_cells('A2:G2')

    # Funding & Valuation
    r=4; ws.cell(row=r,column=1,value='融资与估值历史').font=Font(size=12,bold=True)
    r=5
    f_headers=['时间','轮次','融资额','估值 ($B)','领投方']
    for c,h in enumerate(f_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,5)
    fund_data=[
        ['2023','Series D','$280M','$3.5','Founders Fund, Vy Capital'],
        ['2024初','Secondary','—','$5.0','二级交易'],
        ['2025.06','Series E','$650M','$9.0','ARK Invest, Sequoia, QIA, Valor'],
        ['累计','—','~$1.3B','$9.0','Fidelity, Lightspeed, Thrive等'],
    ]
    for ri,row_d in enumerate(fund_data,6):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # Clinical Milestones
    r=11; ws.cell(row=r,column=1,value='临床里程碑').font=Font(size=12,bold=True)
    r=12
    c_headers=['时间','里程碑','详情']
    for c,h in enumerate(c_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,3)
    clinical=[
        ['2024.01','首例人体植入','Noland Arbaugh (四肢瘫痪) — 用思维控制电脑光标'],
        ['2025.06','第7例植入','累计15,000+小时使用，浏览网页/玩游戏/发帖'],
        ['2025.09','第12例植入','扩展到加拿大、英国、阿联酋临床试验点'],
        ['2025 Q4 (计划)','首个语言皮层植入','解码"默读语音"，帮助失语患者'],
        ['2026 (计划)','电极增至3,000个','精度大幅提升，首例Blindsight视觉植入'],
        ['2029 (目标)','FDA批准Telepathy','美国商业上市，年手术2,000例，收入>$100M'],
        ['2030 (目标)','Blindsight上市','视觉恢复芯片，年手术10,000例，收入>$500M'],
        ['2031 (目标)','5家专科诊所','年植入20,000例，收入≥$1B'],
    ]
    for ri,row_d in enumerate(clinical,13):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            ws.cell(row=ri,column=ci).alignment=WRAP_ALIGN
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # Product Lines
    r=22; ws.cell(row=r,column=1,value='产品线与可触达市场').font=Font(size=12,bold=True)
    r=23
    p_headers=['产品','适应症','全球患者数','预计上市','ASP ($K)','峰值年收入 ($B)']
    for c,h in enumerate(p_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,6)
    products=[
        ['Telepathy','脊髓损伤/ALS','~500万','2029 (FDA)','$50','$5-10'],
        ['Blindsight','失明','~4300万','2030','$80','$10-30'],
        ['Deep','帕金森/震颤','~1000万','2031+','$60','$5-15'],
        ['消费级','增强/娱乐','数亿','2035+ (?)','$1-5','$100+ (远期)'],
    ]
    for ri,row_d in enumerate(products,24):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # BCI Market
    r=29; ws.cell(row=r,column=1,value='BCI 市场规模预测').font=Font(size=12,bold=True)
    r=30
    m_headers=['来源','2029E','2030E','2035E','2040E']
    for c,h in enumerate(m_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,5)
    market_data=[
        ['行业共识 (CAGR 25%)','$7.6B','$12.4B','—','—'],
        ['Bullish (含消费级)','—','$23.7B','—','—'],
        ['McKinsey (医疗+消费)','—','$400B','—','$1,450B'],
        ['Neuralink TAM','$7.6B','$23.7B','$100B+','$500B+'],
    ]
    for ri,row_d in enumerate(market_data,31):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # Scenario
    r=36; ws.cell(row=r,column=1,value='Neuralink 估值情景 (2030E)').font=Font(size=12,bold=True)
    r=37
    n_headers=['指标','Bear','Base','Bull','推导逻辑']
    for c,h in enumerate(n_headers,1): ws.cell(row=r,column=c,value=h)
    style_header_row(ws,r,1,5)
    nlk_scenarios=[
        ['FDA审批','延迟至2031','2029如期','2028加速','医疗器械审批高风险'],
        ['2030手术量','2,000','10,000','25,000','产能爬坡+医生培训'],
        ['2030收入','$100M','$500M','$1.25B','手术量×$50K ASP'],
        ['2030估值','$3B','$15B','$40B','8-10x P/S (高增长MedTech)'],
        ['关键风险','FDA拒绝\n安全性事件','审批延迟\n保险不覆盖','竞品领先\n伦理监管',''],
    ]
    for ri,row_d in enumerate(nlk_scenarios,38):
        for ci,val in enumerate(row_d,1):
            ws.cell(row=ri,column=ci,value=val)
            ws.cell(row=ri,column=ci).border=THIN_BORDER
            ws.cell(row=ri,column=ci).alignment=WRAP_ALIGN
            if ci==1: ws.cell(row=ri,column=ci).font=BLACK_BOLD

    # Key Risk
    r=44; ws.cell(row=r,column=1,value='⚠ 核心风险').font=Font(size=12,bold=True,color='FF0000')
    risks=[
        '• 尚无FDA完全商业批准——所有收入预测取决于监管审批',
        '• 侵入式手术风险——感染、排异、电极退化，任何安全事件可能重创估值',
        '• 保险覆盖不确定——$50K/次的手续如果没有医保覆盖，市场渗透率将极低',
        '• Musk 时间表过于激进——"2026年1000例"vs实际仅12例',
        '• 竞争加速——Synchron (微创支架式)已有10+例，Paradromics完成首例',
    ]
    r=45
    for risk in risks:
        ws.cell(row=r,column=1,value=risk).font=Font(size=10,color='666666')
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
        r+=1

    ws.column_dimensions['A'].width=22
    for c in 'BCDEFG': ws.column_dimensions[c].width=18


if __name__=='__main__':
    main()
