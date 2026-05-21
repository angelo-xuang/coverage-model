#!/usr/bin/env python3
"""
build_excel.py — 生成带场景切换的估值模型 Excel
用法: python build_excel.py --ticker NVDA --input data.json

输入: fetch_financials.py 输出的 JSON 文件（或直接传入参数）
输出: ./out/<ticker>_coverage_model.xlsx

支持:
- 多币种（自动根据公司地区选择 CNY/USD 等）
- 中英文切换（中国运营主体 → 中文，其他 → 英文）
- 假设来源注释列（每个参数旁标明推导依据）
- 精确的 FY→自然年映射说明

依赖: openpyxl
"""

import sys
import json
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ===== 颜色定义 =====
BLUE_FONT = Font(color="0000FF")
BLACK_FONT = Font(color="000000")
GREEN_FONT = Font(color="008000")
GRAY_FONT = Font(color="808080", italic=True)
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BLACK_BOLD = Font(color="000000", bold=True)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
COL_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
OUTPUT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
INPUT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOGGLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NOTE_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

RIGHT_ALIGN = Alignment(horizontal='right')
WRAP_ALIGN = Alignment(horizontal='left', wrap_text=True, vertical='top')


def style_header_row(ws, row, start_col, end_col, fill=HEADER_FILL, font=WHITE_BOLD):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, font=BLACK_FONT, fill=None, num_fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.border = THIN_BORDER
    cell.alignment = RIGHT_ALIGN
    if num_fmt:
        cell.number_format = num_fmt


def get_labels(language='en'):
    """获取中英文标签"""
    if language == 'zh-CN':
        return {
            'sheet_hist': '财务历史',
            'sheet_val': '估值历史',
            'sheet_assumptions': '假设',
            'sheet_forecast': '预测',
            'sheet_sanity': '合理性',
            'sheet_fy_info': '财年信息',
            'label_metric': '指标',
            'active_scenario': '活跃情景',
            'active_value': '活跃值',
            'source_label': '来源/推导依据',
            'revenue': '收入',
            'revenue_growth': '收入增速',
            'cogs': '营业成本',
            'gross_profit': '毛利',
            'gross_margin': '毛利率',
            'rd': '研发费用',
            'rd_ratio': '研发费用率',
            'sga': '销售管理费用',
            'sga_ratio': '销售管理费用率',
            'ebit': 'EBIT',
            'ebit_margin': 'EBIT利润率',
            'net_income': '净利润',
            'net_margin': '净利率',
            'current_mkt_cap': '当前市值',
            'current_price': '当前股价',
            'pe_ttm': 'PE (TTM)',
            'pe_fwd': 'PE (Forward)',
            'shares': '流通股数',
            'wk52_high': '52周高',
            'wk52_low': '52周低',
            'section_revenue': '收入驱动',
            'section_margin': '利润率',
            'section_valuation': '估值',
            'section_rev_forecast': '收入推导',
            'section_profit_forecast': '利润推导',
            'section_val_forecast': '估值',
            'tam_start': 'TAM 起始',
            'tam_cagr_13': 'TAM CAGR Y1-Y3',
            'tam_cagr_46': 'TAM CAGR Y4-Y6',
            'tam_cagr_710': 'TAM CAGR Y7-Y10',
            'reachable': '可触达率',
            'share_start': '起始市占率',
            'share_end': '终年市占率',
            'gm_start': '起始毛利率',
            'gm_end': '终年毛利率',
            'exp_start': '起始费用率',
            'exp_end': '终年费用率',
            'tax_rate': '税率',
            'target_pe': '目标 PE',
            'shares_m': '股份数',
            'tam': 'TAM',
            'sam': '可触达市场',
            'market_share': '市占率',
            'fc_revenue': '收入',
            'fc_rev_growth': '收入增速',
            'fc_gm': '毛利率',
            'fc_gross_profit': '毛利',
            'fc_expense_ratio': '费用率',
            'fc_opex': '运营费用',
            'fc_ebit': 'EBIT',
            'fc_tax_rate': '税率',
            'fc_net_income': '净利润',
            'fc_target_pe': '目标 PE',
            'fc_implied_mkt_cap': '隐含市值',
            'fc_shares': '股份数',
            'fc_implied_price': '隐含股价',
            'fc_premium': 'vs 当前股价溢价',
            'implied_price_label': '隐含股价',
        }
    else:
        return {
            'sheet_hist': 'Financial History',
            'sheet_val': 'Valuation',
            'sheet_assumptions': 'Assumptions',
            'sheet_forecast': 'Forecast',
            'sheet_sanity': 'Sanity Checks',
            'sheet_fy_info': 'Fiscal Year Info',
            'label_metric': 'Metric',
            'active_scenario': 'Active Scenario',
            'active_value': 'Active Value',
            'source_label': 'Source / Derivation',
            'revenue': 'Revenue',
            'revenue_growth': 'Rev Growth',
            'cogs': 'COGS',
            'gross_profit': 'Gross Profit',
            'gross_margin': 'Gross Margin',
            'rd': 'R&D',
            'rd_ratio': 'R&D Ratio',
            'sga': 'SG&A',
            'sga_ratio': 'SG&A Ratio',
            'ebit': 'EBIT',
            'ebit_margin': 'EBIT Margin',
            'net_income': 'Net Income',
            'net_margin': 'Net Margin',
            'current_mkt_cap': 'Current Market Cap',
            'current_price': 'Current Price',
            'pe_ttm': 'PE (TTM)',
            'pe_fwd': 'PE (Forward)',
            'shares': 'Shares Outstanding',
            'wk52_high': '52W High',
            'wk52_low': '52W Low',
            'section_revenue': 'Revenue Drivers',
            'section_margin': 'Profitability',
            'section_valuation': 'Valuation',
            'section_rev_forecast': 'Revenue Build-up',
            'section_profit_forecast': 'Profit Build-up',
            'section_val_forecast': 'Valuation',
            'tam_start': 'TAM Starting',
            'tam_cagr_13': 'TAM CAGR Y1-Y3',
            'tam_cagr_46': 'TAM CAGR Y4-Y6',
            'tam_cagr_710': 'TAM CAGR Y7-Y10',
            'reachable': 'Reachable Rate',
            'share_start': 'Starting Share',
            'share_end': 'Terminal Share',
            'gm_start': 'Starting GM',
            'gm_end': 'Terminal GM',
            'exp_start': 'Starting Exp Ratio',
            'exp_end': 'Terminal Exp Ratio',
            'tax_rate': 'Tax Rate',
            'target_pe': 'Target PE',
            'shares_m': 'Shares (M)',
            'tam': 'TAM',
            'sam': 'Addressable Mkt',
            'market_share': 'Market Share',
            'fc_revenue': 'Revenue',
            'fc_rev_growth': 'Rev Growth',
            'fc_gm': 'Gross Margin',
            'fc_gross_profit': 'Gross Profit',
            'fc_expense_ratio': 'Expense Ratio',
            'fc_opex': 'OpEx',
            'fc_ebit': 'EBIT',
            'fc_tax_rate': 'Tax Rate',
            'fc_net_income': 'Net Income',
            'fc_target_pe': 'Target PE',
            'fc_implied_mkt_cap': 'Implied Mkt Cap',
            'fc_shares': 'Shares (M)',
            'fc_implied_price': 'Implied Price',
            'fc_premium': 'vs Current Premium',
            'implied_price_label': 'Implied Price',
        }


def build_financial_history(ws, data, labels, ccy_sym='$', unit_suffix='M'):
    ws.title = labels['sheet_hist']
    years = sorted(data.keys())
    if not years:
        return

    ws.cell(row=1, column=1, value=labels['label_metric'])
    style_header_row(ws, 1, 1, len(years) + 1)
    for i, y in enumerate(years):
        ws.cell(row=1, column=i + 2, value=y)
        style_header_row(ws, 1, i + 2, i + 2, fill=COL_HEADER_FILL, font=BLACK_BOLD)

    rows_def = [
        (f"{labels['revenue']} ({ccy_sym}{unit_suffix})", "revenue", BLUE_FONT, INPUT_FILL, '#,##0'),
        (labels['revenue_growth'], "revenue_growth", BLACK_FONT, None, '0.0%'),
        (f"{labels['cogs']} ({ccy_sym}{unit_suffix})", "cogs", BLUE_FONT, INPUT_FILL, '#,##0'),
        (f"{labels['gross_profit']} ({ccy_sym}{unit_suffix})", "gross_profit", BLACK_FONT, None, '#,##0'),
        (labels['gross_margin'], "gross_margin", BLACK_FONT, None, '0.0%'),
        (f"{labels['rd']} ({ccy_sym}{unit_suffix})", "rd", BLUE_FONT, INPUT_FILL, '#,##0'),
        (labels['rd_ratio'], "rd_ratio", BLACK_FONT, None, '0.0%'),
        (f"{labels['sga']} ({ccy_sym}{unit_suffix})", "sga", BLUE_FONT, INPUT_FILL, '#,##0'),
        (labels['sga_ratio'], "sga_ratio", BLACK_FONT, None, '0.0%'),
        (f"{labels['ebit']} ({ccy_sym}{unit_suffix})", "ebit", BLACK_FONT, None, '#,##0'),
        (labels['ebit_margin'], "ebit_margin", BLACK_FONT, None, '0.0%'),
        (f"{labels['net_income']} ({ccy_sym}{unit_suffix})", "net_income", BLUE_FONT, INPUT_FILL, '#,##0'),
        (labels['net_margin'], "net_margin", BLACK_FONT, None, '0.0%'),
    ]

    for r_idx, (label, key, font, fill, fmt) in enumerate(rows_def, start=2):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=1).font = BLACK_BOLD
        ws.cell(row=r_idx, column=1).border = THIN_BORDER
        for c_idx, y in enumerate(years, start=2):
            val = data[y].get(key)
            if val is not None:
                ws.cell(row=r_idx, column=c_idx, value=val)
            style_data_cell(ws, r_idx, c_idx, font=font, fill=fill, num_fmt=fmt)

    ws.column_dimensions['A'].width = 26
    for c in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14


def build_valuation_history(ws, val_data, labels, ccy_sym='$'):
    ws.title = labels['sheet_val']
    ws.cell(row=1, column=1, value=labels['label_metric'])
    style_header_row(ws, 1, 1, 1)

    rows = [
        (f"{labels['current_mkt_cap']} ({ccy_sym}B)",
         val_data.get('current_market_cap', 0) / 1e9 if val_data.get('current_market_cap') else None,
         BLUE_FONT, INPUT_FILL, '#,##0.0'),
        (f"{labels['current_price']} ({ccy_sym})",
         val_data.get('current_price'),
         BLUE_FONT, INPUT_FILL, '#,##0.00'),
        (labels['pe_ttm'], val_data.get('current_pe'), BLUE_FONT, INPUT_FILL, '0.0'),
        (labels['pe_fwd'], val_data.get('forward_pe'), BLUE_FONT, INPUT_FILL, '0.0'),
        (f"{labels['shares']} (M)",
         val_data.get('shares_outstanding', 0) / 1e6 if val_data.get('shares_outstanding') else None,
         BLUE_FONT, INPUT_FILL, '#,##0'),
        (f"{labels['wk52_high']} ({ccy_sym})",
         val_data.get('fifty_two_week_high'), BLUE_FONT, INPUT_FILL, '#,##0.00'),
        (f"{labels['wk52_low']} ({ccy_sym})",
         val_data.get('fifty_two_week_low'), BLUE_FONT, INPUT_FILL, '#,##0.00'),
    ]

    for r_idx, (label, val, font, fill, fmt) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=1).font = BLACK_BOLD
        ws.cell(row=r_idx, column=1).border = THIN_BORDER
        if val is not None:
            ws.cell(row=r_idx, column=2, value=val)
        style_data_cell(ws, r_idx, 2, font=font, fill=fill, num_fmt=fmt)

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 14


def build_assumptions(ws, assumptions, labels, ccy_sym='$', unit_suffix='B', sources=None):
    """
    构建假设工作表，含三情景并列 + 下拉切换 + 来源注释列（G列）

    sources dict 的每个 key 对应 assumption_rows 中的 label_key，
    值为一段文字说明该假设是怎么推导出来的。
    """
    ws.title = labels['sheet_assumptions']

    ws.cell(row=1, column=1, value=labels['active_scenario'])
    ws.cell(row=1, column=1).font = WHITE_BOLD
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).border = THIN_BORDER

    ws.cell(row=1, column=2, value=2)
    ws.cell(row=1, column=2).font = BLACK_BOLD
    ws.cell(row=1, column=2).fill = TOGGLE_FILL
    ws.cell(row=1, column=2).border = THIN_BORDER

    dv = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    dv.error = "1=Bear, 2=Base, 3=Bull"
    dv.errorTitle = "Invalid"
    dv.prompt = "1=Bear, 2=Base, 3=Bull"
    dv.promptTitle = "Select Scenario"
    ws.add_data_validation(dv)
    dv.add(ws["B1"])

    # (label, key, fmt, label_key_for_sources)
    # 用 has_seg 检测是否有多分部参数
    has_seg = any(assumptions.get('base', {}).get(k) for k in
                  ['gaming_rev', 'proviz_rev', 'auto_rev', 'oem_rev'])
    assumption_rows = [
        (f"--- {labels['section_revenue']} ---", None, None, None),
        (f"{labels['tam_start']} ({ccy_sym}{unit_suffix})", "tam_start", '#,##0.0', 'tam_start'),
        (labels['tam_cagr_13'], "tam_cagr_1_3", '0.0%', 'tam_cagr_1_3'),
        (labels['tam_cagr_46'], "tam_cagr_4_6", '0.0%', 'tam_cagr_4_6'),
        (labels['tam_cagr_710'], "tam_cagr_7_10", '0.0%', 'tam_cagr_7_10'),
        (labels['reachable'], "reachable_rate", '0.0%', 'reachable_rate'),
        (labels['share_start'], "share_start", '0.0%', 'share_start'),
        (labels['share_end'], "share_end", '0.0%', 'share_end'),
    ]
    if has_seg:
        assumption_rows += [
            (f"--- {labels.get('section_gaming', 'Gaming')} ---", None, None, None),
            (f"Gaming 起始收入 ({ccy_sym}{unit_suffix})", "gaming_rev", '#,##0.0', 'gaming_rev'),
            ("Gaming CAGR Y1-Y3", "gaming_cagr_1_3", '0.0%', 'gaming_cagr_1_3'),
            ("Gaming CAGR Y4-Y6", "gaming_cagr_4_6", '0.0%', 'gaming_cagr_4_6'),
            ("Gaming CAGR Y7-Y10", "gaming_cagr_7_10", '0.0%', 'gaming_cagr_7_10'),
            (f"--- {labels.get('section_proviz', 'ProViz')} ---", None, None, None),
            (f"ProViz 起始收入 ({ccy_sym}{unit_suffix})", "proviz_rev", '#,##0.0', 'proviz_rev'),
            ("ProViz CAGR Y1-Y3", "proviz_cagr_1_3", '0.0%', 'proviz_cagr_1_3'),
            ("ProViz CAGR Y4-Y6", "proviz_cagr_4_6", '0.0%', 'proviz_cagr_4_6'),
            ("ProViz CAGR Y7-Y10", "proviz_cagr_7_10", '0.0%', 'proviz_cagr_7_10'),
            (f"--- {labels.get('section_auto', 'Auto')} ---", None, None, None),
            (f"Auto 起始收入 ({ccy_sym}{unit_suffix})", "auto_rev", '#,##0.0', 'auto_rev'),
            ("Auto CAGR Y1-Y3", "auto_cagr_1_3", '0.0%', 'auto_cagr_1_3'),
            ("Auto CAGR Y4-Y6", "auto_cagr_4_6", '0.0%', 'auto_cagr_4_6'),
            ("Auto CAGR Y7-Y10", "auto_cagr_7_10", '0.0%', 'auto_cagr_7_10'),
            (f"--- {labels.get('section_oem', 'OEM')} ---", None, None, None),
            (f"OEM 收入 ({ccy_sym}{unit_suffix})", "oem_rev", '#,##0.0', 'oem_rev'),
        ]
    assumption_rows += [
        (f"--- {labels['section_margin']} ---", None, None, None),
        (labels['gm_start'], "gm_start", '0.0%', 'gm_start'),
        (labels['gm_end'], "gm_end", '0.0%', 'gm_end'),
        (labels['exp_start'], "expense_start", '0.0%', 'expense_start'),
        (labels['exp_end'], "expense_end", '0.0%', 'expense_end'),
        (labels['tax_rate'], "tax_rate", '0.0%', 'tax_rate'),
        (f"--- {labels['section_valuation']} ---", None, None, None),
        (f"{labels['target_pe']} 起始", "pe_start", '0.0', 'pe_start'),
        (f"{labels['target_pe']} 终年", "pe_end", '0.0', 'pe_end'),
        (f"{labels['shares_m']} (M)", "shares_m", '#,##0', 'shares_m'),
    ]

    # 列标题: A=假设项 B=下拉 C=Bear D=Base E=Bull F=活跃值 G=来源
    ws.cell(row=2, column=1, value=labels['label_metric'])
    ws.cell(row=2, column=3, value="Bear")
    ws.cell(row=2, column=4, value="Base")
    ws.cell(row=2, column=5, value="Bull")
    ws.cell(row=2, column=6, value=labels['active_value'])
    ws.cell(row=2, column=7, value=labels['source_label'])
    style_header_row(ws, 2, 1, 7)
    style_header_row(ws, 2, 3, 5, fill=COL_HEADER_FILL, font=BLACK_BOLD)
    style_header_row(ws, 2, 6, 6, fill=OUTPUT_FILL, font=BLACK_BOLD)
    style_header_row(ws, 2, 7, 7,
                     fill=PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
                     font=Font(color="666666", bold=True))

    for r_idx, (label, key, fmt, label_key) in enumerate(assumption_rows, start=3):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=1).font = BLACK_BOLD
        ws.cell(row=r_idx, column=1).border = THIN_BORDER

        if key is None:
            continue

        for c_idx, scenario in enumerate(['bear', 'base', 'bull'], start=3):
            val = assumptions.get(scenario, {}).get(key)
            if val is not None:
                ws.cell(row=r_idx, column=c_idx, value=val)
            style_data_cell(ws, r_idx, c_idx, font=BLUE_FONT, fill=INPUT_FILL, num_fmt=fmt)

        ws.cell(row=r_idx, column=6, value=f'=INDEX(C{r_idx}:E{r_idx},$B$1)')
        style_data_cell(ws, r_idx, 6, font=GREEN_FONT, fill=OUTPUT_FILL, num_fmt=fmt)

        # G列: 来源/推导依据
        source_text = ''
        if sources and label_key in sources:
            source_text = sources[label_key]
        ws.cell(row=r_idx, column=7, value=source_text)
        ws.cell(row=r_idx, column=7).font = GRAY_FONT
        ws.cell(row=r_idx, column=7).fill = NOTE_FILL
        ws.cell(row=r_idx, column=7).border = THIN_BORDER
        ws.cell(row=r_idx, column=7).alignment = WRAP_ALIGN

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 8
    for c in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 14
    ws.column_dimensions['G'].width = 55


def build_projections(ws, assumptions, hist_data, labels, ccy_sym='$',
                       unit_suffix='B', projection_years=10, current_price=None):
    """预测工作表 — 参数行[绿]嵌入计算链，首列=最新实际年[B列/蓝]

    布局: 参数[绿]在上 → 计算[黑]在下，绿色=跨表引用假设sheet F列
    行号严格递增(2-56)，无重叠
    """
    ws.title = labels['sheet_forecast']
    VAL = current_price or 100
    sn = labels['sheet_assumptions']
    N = projection_years
    hist_years = sorted(hist_data.keys())
    ly_str = hist_years[-1] if hist_years else '2025A'
    ly = int(ly_str.replace('A', ''))
    yrs = [ly_str] + [f"{ly + i + 1}E" for i in range(N)]
    TC = len(yrs)  # = 1 历史 + N 预测 = 11

    ws.cell(row=1, column=1, value=labels['label_metric'])
    for i, y in enumerate(yrs):
        ws.cell(row=1, column=i + 2, value=y)
    style_header_row(ws, 1, 1, TC + 1)
    style_header_row(ws, 1, 2, TC + 1, fill=COL_HEADER_FILL, font=BLACK_BOLD)

    hist = hist_data.get(ly_str, {})
    h_rev = hist.get('revenue', 0)
    h_rev_B = h_rev / 1000 if h_rev > 1e6 else h_rev
    h_gm = hist.get('gross_margin', 0)
    h_gr = hist.get('revenue_growth', 0)
    h_ni = hist.get('net_income', 0)
    h_ni_B = h_ni / 1000 if h_ni > 1e6 else h_ni

    # 检测是否有分部参数 (Gaming等)
    base_asm = assumptions.get('base', {})
    has_seg = any(base_asm.get(k) for k in
                  ['gaming_rev', 'proviz_rev', 'auto_rev', 'oem_rev'])

    # 动态计算假设行号: DC(3-10) → [segments] → Margin → Valuation
    _r = 4
    AR = {'tam_start': _r, 'tam_cagr_1_3': _r+1, 'tam_cagr_4_6': _r+2, 'tam_cagr_7_10': _r+3}
    _r += 4  # 8
    AR['reachable_rate'] = _r; AR['share_start'] = _r+1; AR['share_end'] = _r+2
    _r = _r + 4  # 12 (or more if segments)
    if has_seg:
        for seg in ['gaming', 'proviz', 'auto']:
            AR[f'{seg}_rev'] = _r; _r += 1
            AR[f'{seg}_cagr_1_3'] = _r; _r += 1
            AR[f'{seg}_cagr_4_6'] = _r; _r += 1
            AR[f'{seg}_cagr_7_10'] = _r; _r += 1
        AR['oem_rev'] = _r; _r += 1
    AR['gm_start'] = _r; AR['gm_end'] = _r+1
    AR['expense_start'] = _r+2; AR['expense_end'] = _r+3
    AR['tax_rate'] = _r+4
    _r += 6
    AR['pe_start'] = _r; AR['pe_end'] = _r+1
    AR['shares_m'] = _r+2

    def asum(k):
        return f"'{sn}'!$F${AR[k]}"

    def G(r, label, key, fmt='0.0%'):
        """绿色参数行"""
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = BLACK_BOLD; ws.cell(row=r, column=1).border = THIN_BORDER
        for i in range(TC):
            ws.cell(row=r, column=i + 2, value=f"={asum(key)}")
            style_data_cell(ws, r, i + 2, font=GREEN_FONT, num_fmt=fmt)

    def S(r, label):
        """节标题"""
        ws.cell(row=r, column=1, value=f"--- {label} ---")
        ws.cell(row=r, column=1).font = WHITE_BOLD
        ws.cell(row=r, column=1).fill = HEADER_FILL
        for c in range(2, TC + 2):
            ws.cell(row=r, column=c).fill = HEADER_FILL

    def L(r, text):
        """黑体标签"""
        ws.cell(row=r, column=1, value=text)
        ws.cell(row=r, column=1).font = BLACK_BOLD; ws.cell(row=r, column=1).border = THIN_BORDER

    def B(r, col, val, fmt='#,##0.0'):
        """蓝底历史值"""
        if val is not None:
            ws.cell(row=r, column=col, value=val)
        style_data_cell(ws, r, col, font=BLUE_FONT, fill=INPUT_FILL, num_fmt=fmt)

    def O(r, col, formula, fmt='#,##0.0'):
        """黑字输出(蓝底)"""
        ws.cell(row=r, column=col, value=formula)
        style_data_cell(ws, r, col, font=BLACK_FONT, fill=OUTPUT_FILL, num_fmt=fmt)

    def K(r, col, formula, fmt='#,##0.0'):
        """黑字计算(无底)"""
        ws.cell(row=r, column=col, value=formula)
        style_data_cell(ws, r, col, font=BLACK_FONT, num_fmt=fmt)

    # ============================================
    # 行号 2-13: 收入推导 (DC TAM模型)
    # ============================================
    S(2, labels['section_rev_forecast'])
    G(3, f"TAM起始 ({ccy_sym}{unit_suffix})", 'tam_start', '#,##0.0')
    G(4, "TAM CAGR Y1-Y3", 'tam_cagr_1_3')
    G(5, "TAM CAGR Y4-Y6", 'tam_cagr_4_6')
    G(6, "TAM CAGR Y7-Y10", 'tam_cagr_7_10')

    L(7, f"TAM ({ccy_sym}{unit_suffix})")
    B(7, 2, f"={asum('tam_start')}", '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        r_cagr = 4 if i <= 3 else (5 if i <= 6 else 6)
        if i == 1:
            K(7, col, f"={c}3*(1+{c}{r_cagr})")
        else:
            pc = get_column_letter(col - 1)
            K(7, col, f"={pc}7*(1+{c}{r_cagr})")

    G(8, labels['reachable'], 'reachable_rate')

    L(9, f"{labels['sam']} ({ccy_sym}{unit_suffix})")
    for i in range(0, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(9, col, f"={c}7*{c}8")

    G(10, labels['share_start'], 'share_start')
    G(11, labels['share_end'], 'share_end')

    L(12, labels['market_share'])
    for i in range(0, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(12, col, f"={c}10" if i == 0 else f"={c}10+({c}11-{c}10)*{i}/{N}", '0.0%')

    L(13, f"DC 收入 ({ccy_sym}{unit_suffix})")
    B(13, 2, h_rev_B, '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        O(13, col, f"={c}9*{c}12")

    # ============================================
    # 分部收入 (仅 has_seg 时渲染) + 总收入
    # ============================================
    _r = 14  # 当前渲染行
    seg_rows = {}  # 记录每个分部的收入行号用于汇总
    if has_seg:
        for seg, seg_label in [('gaming', 'Gaming'), ('proviz', 'ProViz'),
                                ('auto', 'Auto')]:
            S(_r, labels.get(f'section_{seg.lower()}', seg_label)); _r += 1
            G(_r, f"{seg_label} 起始 ({ccy_sym}{unit_suffix})", f'{seg}_rev', '#,##0.0'); _r += 1
            G(_r, f"{seg_label} CAGR Y1-Y3", f'{seg}_cagr_1_3'); _r += 1
            G(_r, f"{seg_label} CAGR Y4-Y6", f'{seg}_cagr_4_6'); _r += 1
            G(_r, f"{seg_label} CAGR Y7-Y10", f'{seg}_cagr_7_10'); _r += 1
            r_rev = _r
            L(_r, f"{seg_label} 收入 ({ccy_sym}{unit_suffix})"); _r += 1
            seg_rows[seg] = r_rev
            for i in range(1, N + 1):
                col = i + 2; c = get_column_letter(col); pc = get_column_letter(col - 1)
                r_cagr = r_rev - 3 if i <= 3 else (r_rev - 2 if i <= 6 else r_rev - 1)
                K(r_rev, col,
                  f"={c}{r_rev-4}*(1+{c}{r_cagr})" if i == 1 else f"={pc}{r_rev}*(1+{c}{r_cagr})")

        # OEM
        S(_r, labels.get('section_oem', 'OEM')); _r += 1
        r_oem = _r
        L(_r, f"OEM 收入 ({ccy_sym}{unit_suffix})"); _r += 1
        for i in range(1, N + 1):
            col = i + 2
            ws.cell(row=r_oem, column=col, value=f"={asum('oem_rev')}")
            style_data_cell(ws, r_oem, col, font=GREEN_FONT, num_fmt='#,##0.0')

    # 总收入 (r_rev_total)
    S(_r, labels.get('total_revenue', 'Total Revenue')); _r += 1
    r_total = _r
    L(_r, f"{labels.get('total_revenue', 'Total')} ({ccy_sym}{unit_suffix})"); _r += 1
    B(r_total, 2, h_rev_B, '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        if has_seg:
            parts = '+'.join(f"{c}{seg_rows[s]}" for s in ['gaming', 'proviz', 'auto'] + ['oem'] if s in seg_rows)
            # OEM row
            parts = f"{c}13+" + '+'.join(f"{c}{seg_rows[s]}" for s in ['gaming', 'proviz', 'auto'])
            parts += f"+{c}{r_oem}"
            O(r_total, col, parts)
        else:
            O(r_total, col, f"={c}13")

    r_growth = _r
    L(_r, f"总收入 增速"); _r += 1
    B(r_growth, 2, h_gr, '0.0%')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col); pc = get_column_letter(col - 1)
        K(r_growth, col, f"=({c}{r_total}-{pc}{r_total})/{pc}{r_total}", '0.0%')

    # ============================================
    # 利润推导
    # ============================================
    S(_r, labels['section_profit_forecast']); _r += 1
    r_gm_start = _r; G(_r, labels['gm_start'], 'gm_start'); _r += 1
    r_gm_end = _r; G(_r, labels['gm_end'], 'gm_end'); _r += 1

    r_gm = _r; L(_r, labels['fc_gm']); _r += 1
    B(r_gm, 2, h_gm, '0.0%')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_gm, col, f"={c}{r_gm_start}+({c}{r_gm_end}-{c}{r_gm_start})*{i - 1}/{N - 1}", '0.0%')

    r_gp = _r; L(_r, f"{labels['fc_gross_profit']} ({ccy_sym}{unit_suffix})"); _r += 1
    if h_rev_B and h_gm:
        B(r_gp, 2, h_rev_B * h_gm, '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_gp, col, f"={c}{r_total}*{c}{r_gm}")

    r_exp_s = _r; G(_r, labels['exp_start'], 'expense_start'); _r += 1
    r_exp_e = _r; G(_r, labels['exp_end'], 'expense_end'); _r += 1

    r_exp = _r; L(_r, labels['fc_expense_ratio']); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_exp, col, f"={c}{r_exp_s}+({c}{r_exp_e}-{c}{r_exp_s})*{i - 1}/{N - 1}", '0.0%')

    r_opex = _r; L(_r, f"{labels['fc_opex']} ({ccy_sym}{unit_suffix})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_opex, col, f"={c}{r_total}*{c}{r_exp}")

    r_ebit = _r; L(_r, f"{labels['fc_ebit']} ({ccy_sym}{unit_suffix})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_ebit, col, f"={c}{r_gp}-{c}{r_opex}")

    r_tax = _r; G(_r, labels['tax_rate'], 'tax_rate'); _r += 1

    r_ni = _r; L(_r, f"{labels['fc_net_income']} ({ccy_sym}{unit_suffix})"); _r += 1
    B(r_ni, 2, h_ni_B, '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        O(r_ni, col, f"={c}{r_ebit}*(1-{c}{r_tax})")

    # ============================================
    # 估值
    # ============================================
    S(_r, labels['section_val_forecast']); _r += 1
    r_pe_s = _r; G(_r, f"{labels['target_pe']} 起始", 'pe_start', '0.0'); _r += 1
    r_pe_e = _r; G(_r, f"{labels['target_pe']} 终年", 'pe_end', '0.0'); _r += 1

    r_pe = _r; L(_r, labels['fc_target_pe']); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_pe, col, f"={c}{r_pe_s}+({c}{r_pe_e}-{c}{r_pe_s})*{i - 1}/{N - 1}", '0.0')

    r_mkt = _r; L(_r, f"{labels['fc_implied_mkt_cap']} ({ccy_sym}{unit_suffix})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        O(r_mkt, col, f"={c}{r_ni}*{c}{r_pe}")

    r_sh = _r; G(_r, labels['fc_shares'], 'shares_m', '#,##0'); _r += 1

    r_price = _r; L(_r, f"{labels['implied_price_label']} ({ccy_sym})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        O(r_price, col, f"={c}{r_mkt}*1000/{c}{r_sh}", '#,##0.00')

    r_prem = _r; L(_r, labels['fc_premium']); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_prem, col, f"=({c}{r_price}-{VAL})/{VAL}", '0.0%')

    ws.column_dimensions['A'].width = 28
    for c in range(2, TC + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14


def build_segment_assumptions(ws, seg_asm, labels, ccy_sym='$', unit_suffix='B', sources=None):
    """分部模型假设表 — 每分部独立TAM×市占率 + 公司级利润率/估值参数"""
    ws.title = labels['sheet_assumptions']

    ws.cell(row=1, column=1, value=labels['active_scenario'])
    ws.cell(row=1, column=1).font = WHITE_BOLD; ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).border = THIN_BORDER
    ws.cell(row=1, column=2, value=2)
    ws.cell(row=1, column=2).font = BLACK_BOLD; ws.cell(row=1, column=2).fill = TOGGLE_FILL
    ws.cell(row=1, column=2).border = THIN_BORDER

    dv = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    dv.error = "1=Bear, 2=Base, 3=Bull"; dv.prompt = "1=Bear, 2=Base, 3=Bull"
    ws.add_data_validation(dv); dv.add(ws["B1"])

    ws.cell(row=2, column=1, value=labels['label_metric'])
    ws.cell(row=2, column=3, value="Bear"); ws.cell(row=2, column=4, value="Base")
    ws.cell(row=2, column=5, value="Bull"); ws.cell(row=2, column=6, value=labels['active_value'])
    ws.cell(row=2, column=7, value=labels['source_label'])
    style_header_row(ws, 2, 1, 7)
    style_header_row(ws, 2, 3, 5, fill=COL_HEADER_FILL, font=BLACK_BOLD)
    style_header_row(ws, 2, 6, 6, fill=OUTPUT_FILL, font=BLACK_BOLD)
    style_header_row(ws, 2, 7, 7, fill=PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
                     font=Font(color="666666", bold=True))

    # 分部定义: (display_name, segment_key, params_list)
    # params_list: [(label_suffix, data_key, fmt, source_key), ...]
    seg_defs = [
        ("Search & Other", "search", [
            (f"TAM起始 ({ccy_sym}{unit_suffix})", "tam_2025", '#,##0.0', 'search_tam'),
            ("TAM CAGR Y1-Y3", "tam_cagr_1_3", '0.0%', 'search_cagr_13'),
            ("TAM CAGR Y4-Y6", "tam_cagr_4_6", '0.0%', 'search_cagr_46'),
            ("TAM CAGR Y7-Y10", "tam_cagr_7_10", '0.0%', 'search_cagr_710'),
            ("市占率 起始", "share_2025", '0.0%', 'search_share_start'),
            ("市占率 终年", "share_2035", '0.0%', 'search_share_end'),
        ]),
        ("YouTube Ads", "youtube", [
            (f"TAM起始 ({ccy_sym}{unit_suffix})", "tam_2025", '#,##0.0', 'youtube_tam'),
            ("TAM CAGR Y1-Y3", "tam_cagr_1_3", '0.0%', 'youtube_cagr_13'),
            ("TAM CAGR Y4-Y6", "tam_cagr_4_6", '0.0%', 'youtube_cagr_46'),
            ("TAM CAGR Y7-Y10", "tam_cagr_7_10", '0.0%', 'youtube_cagr_710'),
            ("市占率 起始", "share_2025", '0.0%', 'youtube_share_start'),
            ("市占率 终年", "share_2035", '0.0%', 'youtube_share_end'),
        ]),
        ("Google Cloud", "cloud", [
            (f"TAM起始 ({ccy_sym}{unit_suffix})", "tam_2025", '#,##0.0', 'cloud_tam'),
            ("TAM CAGR Y1-Y3", "tam_cagr_1_3", '0.0%', 'cloud_cagr_13'),
            ("TAM CAGR Y4-Y6", "tam_cagr_4_6", '0.0%', 'cloud_cagr_46'),
            ("TAM CAGR Y7-Y10", "tam_cagr_7_10", '0.0%', 'cloud_cagr_710'),
            ("市占率 起始", "share_2025", '0.0%', 'cloud_share_start'),
            ("市占率 终年", "share_2035", '0.0%', 'cloud_share_end'),
        ]),
        ("Google Network", "network", [
            (f"起始收入 ({ccy_sym}{unit_suffix})", "revenue_2025", '#,##0.0', 'network_rev'),
            ("年衰减率", "annual_decline", '0.0%', 'network_decline'),
        ]),
        ("Subscriptions", "subscriptions", [
            (f"TAM起始 ({ccy_sym}{unit_suffix})", "tam_2025", '#,##0.0', 'subs_tam'),
            ("TAM CAGR Y1-Y3", "tam_cagr_1_3", '0.0%', 'subs_cagr_13'),
            ("TAM CAGR Y4-Y6", "tam_cagr_4_6", '0.0%', 'subs_cagr_46'),
            ("TAM CAGR Y7-Y10", "tam_cagr_7_10", '0.0%', 'subs_cagr_710'),
            ("市占率 起始", "share_2025", '0.0%', 'subs_share_start'),
            ("市占率 终年", "share_2035", '0.0%', 'subs_share_end'),
        ]),
    ]

    # 公司级参数
    co_params = [
        (f"--- {labels['section_margin']} ---", None, None, None, None),
        (labels['gm_start'], None, "gm_start", '0.0%', 'gm_start'),
        (labels['gm_end'], None, "gm_end", '0.0%', 'gm_end'),
        (labels['exp_start'], None, "expense_start", '0.0%', 'expense_start'),
        (labels['exp_end'], None, "expense_end", '0.0%', 'expense_end'),
        (labels['tax_rate'], None, "tax_rate", '0.0%', 'tax_rate'),
        (f"--- {labels['section_valuation']} ---", None, None, None, None),
        (f"{labels['target_pe']} 起始", None, "pe_start", '0.0', 'pe_start'),
        (f"{labels['target_pe']} 终年", None, "pe_end", '0.0', 'pe_end'),
        (f"{labels['shares_m']} (M)", None, "shares_m", '#,##0', 'shares_m'),
    ]

    _r = 3
    # 分部参数
    for seg_name, seg_key, params in seg_defs:
        ws.cell(row=_r, column=1, value=f"--- {seg_name} ---")
        ws.cell(row=_r, column=1).font = BLACK_BOLD
        ws.cell(row=_r, column=1).border = THIN_BORDER
        _r += 1
        for label_suffix, dk, fmt, sk in params:
            ws.cell(row=_r, column=1, value=label_suffix)
            ws.cell(row=_r, column=1).font = BLACK_BOLD
            ws.cell(row=_r, column=1).border = THIN_BORDER
            for ci, sc in enumerate(['bear', 'base', 'bull'], start=3):
                seg_data = seg_asm.get(sc, {}).get(seg_key, {})
                val = seg_data.get(dk) if seg_data else None
                if val is not None:
                    ws.cell(row=_r, column=ci, value=val)
                style_data_cell(ws, _r, ci, font=BLUE_FONT, fill=INPUT_FILL, num_fmt=fmt)
            ws.cell(row=_r, column=6, value=f'=INDEX(C{_r}:E{_r},$B$1)')
            style_data_cell(ws, _r, 6, font=GREEN_FONT, fill=OUTPUT_FILL, num_fmt=fmt)
            st = ''
            if sources and sk in sources:
                st = sources[sk]
            ws.cell(row=_r, column=7, value=st)
            ws.cell(row=_r, column=7).font = GRAY_FONT
            ws.cell(row=_r, column=7).fill = NOTE_FILL
            ws.cell(row=_r, column=7).border = THIN_BORDER
            ws.cell(row=_r, column=7).alignment = WRAP_ALIGN
            _r += 1

    # 公司级参数
    for label, seg_key, dk, fmt, sk in co_params:
        if seg_key is None and dk is None:
            ws.cell(row=_r, column=1, value=label)
            ws.cell(row=_r, column=1).font = BLACK_BOLD
            ws.cell(row=_r, column=1).border = THIN_BORDER
            _r += 1
            continue
        ws.cell(row=_r, column=1, value=label)
        ws.cell(row=_r, column=1).font = BLACK_BOLD
        ws.cell(row=_r, column=1).border = THIN_BORDER
        for ci, sc in enumerate(['bear', 'base', 'bull'], start=3):
            val = seg_asm.get(sc, {}).get(dk)
            if val is not None:
                ws.cell(row=_r, column=ci, value=val)
            style_data_cell(ws, _r, ci, font=BLUE_FONT, fill=INPUT_FILL, num_fmt=fmt)
        ws.cell(row=_r, column=6, value=f'=INDEX(C{_r}:E{_r},$B$1)')
        style_data_cell(ws, _r, 6, font=GREEN_FONT, fill=OUTPUT_FILL, num_fmt=fmt)
        st = ''
        if sources and sk in sources:
            st = sources[sk]
        ws.cell(row=_r, column=7, value=st)
        ws.cell(row=_r, column=7).font = GRAY_FONT
        ws.cell(row=_r, column=7).fill = NOTE_FILL
        ws.cell(row=_r, column=7).border = THIN_BORDER
        ws.cell(row=_r, column=7).alignment = WRAP_ALIGN
        _r += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 8
    for c in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 14
    ws.column_dimensions['G'].width = 55

    # 返回各分部参数的行号映射
    return _r


def build_segment_projections(ws, seg_asm, hist_data, labels, ccy_sym='$',
                               unit_suffix='B', projection_years=10, current_price=None):
    """分部模型预测表 — 每分部独立 TAM×市占率 → 总收入 → 利润 → 估值"""
    ws.title = labels['sheet_forecast']
    VAL = current_price or 100
    sn = labels['sheet_assumptions']
    N = projection_years
    hist_years = sorted(hist_data.keys())
    ly_str = hist_years[-1] if hist_years else '2025A'
    ly = int(ly_str.replace('A', ''))
    yrs = [ly_str] + [f"{ly + i + 1}E" for i in range(N)]
    TC = len(yrs)

    ws.cell(row=1, column=1, value=labels['label_metric'])
    for i, y in enumerate(yrs):
        ws.cell(row=1, column=i + 2, value=y)
    style_header_row(ws, 1, 1, TC + 1)
    style_header_row(ws, 1, 2, TC + 1, fill=COL_HEADER_FILL, font=BLACK_BOLD)

    hist = hist_data.get(ly_str, {})
    h_rev = hist.get('revenue', 0)
    h_rev_B = h_rev / 1000 if h_rev > 1e6 else h_rev
    h_gm = hist.get('gross_margin', 0)
    h_gr = hist.get('revenue_growth', 0)
    h_ni = hist.get('net_income', 0)
    h_ni_B = h_ni / 1000 if h_ni > 1e6 else h_ni

    # 短辅助
    def S(r, label):
        ws.cell(row=r, column=1, value=f"--- {label} ---")
        ws.cell(row=r, column=1).font = WHITE_BOLD; ws.cell(row=r, column=1).fill = HEADER_FILL
        for c in range(2, TC + 2):
            ws.cell(row=r, column=c).fill = HEADER_FILL

    def L(r, text):
        ws.cell(row=r, column=1, value=text)
        ws.cell(row=r, column=1).font = BLACK_BOLD; ws.cell(row=r, column=1).border = THIN_BORDER

    def B(r, col, val, fmt='#,##0.0'):
        if val is not None:
            ws.cell(row=r, column=col, value=val)
        style_data_cell(ws, r, col, font=BLUE_FONT, fill=INPUT_FILL, num_fmt=fmt)

    def O(r, col, formula, fmt='#,##0.0'):
        ws.cell(row=r, column=col, value=formula)
        style_data_cell(ws, r, col, font=BLACK_FONT, fill=OUTPUT_FILL, num_fmt=fmt)

    def K(r, col, formula, fmt='#,##0.0'):
        ws.cell(row=r, column=col, value=formula)
        style_data_cell(ws, r, col, font=BLACK_FONT, num_fmt=fmt)

    # 计算假设行的函数: 需要知道每个分部参数的起始行
    # 分部参数布局: 每个分部 header(1) + params(N), 公司级参数跟在后面
    # 我们基于 seg_defs 计算
    seg_specs = [
        ('search', 'Search & Other', True),   # TAM-based
        ('youtube', 'YouTube Ads', True),
        ('cloud', 'Google Cloud', True),
        ('network', 'Google Network', False),  # declining
        ('subscriptions', 'Subscriptions', True),
    ]

    # 假设行行号: header_row(3) → search(header4, params5-10) → youtube(header11, params12-17) → ...
    arow = {}  # segment → {param_key → assumption_row}
    _ar = 4
    for skey, _, is_tam in seg_specs:
        _ar += 1  # skip header
        if is_tam:
            arow[skey] = {
                'tam_start': _ar, 'tam_cagr_1_3': _ar+1, 'tam_cagr_4_6': _ar+2,
                'tam_cagr_7_10': _ar+3, 'share_start': _ar+4, 'share_end': _ar+5,
            }
            _ar += 6
        else:
            arow[skey] = {'revenue_start': _ar, 'annual_decline': _ar+1}
            _ar += 2

    # 公司级参数: (skip section header rows)
    _ar += 1  # margin section header
    co_ar = {
        'gm_start': _ar, 'gm_end': _ar+1, 'expense_start': _ar+2, 'expense_end': _ar+3,
        'tax_rate': _ar+4,
    }
    _ar += 6  # valuation section header
    co_ar['pe_start'] = _ar; co_ar['pe_end'] = _ar+1; co_ar['shares_m'] = _ar+2

    def asum_co(key):
        return f"'{sn}'!$F${co_ar[key]}"

    def asum_seg(seg_key, param_key):
        return f"'{sn}'!$F${arow[seg_key][param_key]}"

    # ============================================
    # 各分部收入推导
    # ============================================
    _r = 2
    seg_rev_rows = {}

    for skey, sname, is_tam in seg_specs:
        S(_r, sname); _r += 1  # section header

        if is_tam:
            # TAM参数行(绿)
            ws.cell(row=_r, column=1, value=f"TAM 起始 ({ccy_sym}{unit_suffix})")
            ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
            for i in range(TC):
                ws.cell(row=_r, column=i + 2, value=f"={asum_seg(skey, 'tam_start')}")
                style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='#,##0.0')
            r_ts = _r; _r += 1

            for lbl, pk in [("TAM CAGR Y1-Y3", 'tam_cagr_1_3'),
                             ("TAM CAGR Y4-Y6", 'tam_cagr_4_6'),
                             ("TAM CAGR Y7-Y10", 'tam_cagr_7_10')]:
                ws.cell(row=_r, column=1, value=lbl)
                ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
                for i in range(TC):
                    ws.cell(row=_r, column=i + 2, value=f"={asum_seg(skey, pk)}")
                    style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
                _r += 1

            # TAM 计算
            r_tam = _r
            L(_r, f"TAM ({ccy_sym}{unit_suffix})"); _r += 1
            for i in range(1, N + 1):
                col = i + 2; c = get_column_letter(col)
                r_c = r_ts + 1 if i <= 3 else (r_ts + 2 if i <= 6 else r_ts + 3)
                if i == 1:
                    K(r_tam, col, f"={c}{r_ts}*(1+{c}{r_c})")
                else:
                    pc = get_column_letter(col - 1)
                    K(r_tam, col, f"={pc}{r_tam}*(1+{c}{r_c})")

            # 市占率参数(绿)
            ws.cell(row=_r, column=1, value="市占率 起始")
            ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
            for i in range(TC):
                ws.cell(row=_r, column=i + 2, value=f"={asum_seg(skey, 'share_start')}")
                style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
            r_ss = _r; _r += 1

            ws.cell(row=_r, column=1, value="市占率 终年")
            ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
            for i in range(TC):
                ws.cell(row=_r, column=i + 2, value=f"={asum_seg(skey, 'share_end')}")
                style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
            r_se = _r; _r += 1

            # 市占率 线性插值
            r_sh = _r
            L(_r, "市占率"); _r += 1
            for i in range(0, N + 1):
                col = i + 2; c = get_column_letter(col)
                if i == 0:
                    K(r_sh, col, f"={c}{r_ss}", '0.0%')
                else:
                    K(r_sh, col, f"={c}{r_ss}+({c}{r_se}-{c}{r_ss})*{i}/{N}", '0.0%')

            # 分部收入 = TAM × 市占率
            r_srev = _r
            L(_r, f"{sname} 收入 ({ccy_sym}{unit_suffix})"); _r += 1
            for i in range(1, N + 1):
                col = i + 2; c = get_column_letter(col)
                O(r_srev, col, f"={c}{r_tam}*{c}{r_sh}")
            seg_rev_rows[skey] = r_srev

        else:
            # Network: declining
            ws.cell(row=_r, column=1, value=f"起始收入 ({ccy_sym}{unit_suffix})")
            ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
            for i in range(TC):
                ws.cell(row=_r, column=i + 2, value=f"={asum_seg(skey, 'revenue_start')}")
                style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='#,##0.0')
            r_ns = _r; _r += 1

            ws.cell(row=_r, column=1, value="年衰减率")
            ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
            for i in range(TC):
                ws.cell(row=_r, column=i + 2, value=f"={asum_seg(skey, 'annual_decline')}")
                style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
            r_nd = _r; _r += 1

            r_nrev = _r
            L(_r, f"{sname} 收入 ({ccy_sym}{unit_suffix})"); _r += 1
            for i in range(1, N + 1):
                col = i + 2; c = get_column_letter(col)
                if i == 1:
                    K(r_nrev, col, f"={c}{r_ns}*(1+{c}{r_nd})")
                else:
                    pc = get_column_letter(col - 1)
                    K(r_nrev, col, f"={pc}{r_nrev}*(1+{c}{r_nd})")
            seg_rev_rows[skey] = r_nrev

    # ============================================
    # 总收入
    # ============================================
    S(_r, "Total Revenue"); _r += 1
    r_total = _r
    L(_r, f"总收入 ({ccy_sym}{unit_suffix})"); _r += 1
    B(r_total, 2, h_rev_B, '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        parts = '+'.join(f"{c}{seg_rev_rows[s]}" for s, _, _ in seg_specs)
        O(r_total, col, parts)

    r_growth = _r
    L(_r, "总收入 增速"); _r += 1
    B(r_growth, 2, h_gr, '0.0%')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col); pc = get_column_letter(col - 1)
        K(r_growth, col, f"=({c}{r_total}-{pc}{r_total})/{pc}{r_total}", '0.0%')

    # ============================================
    # 利润推导 (同简单模型, 使用动态行号)
    # ============================================
    S(_r, labels['section_profit_forecast']); _r += 1
    r_gm_s = _r
    ws.cell(row=_r, column=1, value=labels['gm_start'])
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('gm_start')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
    _r += 1

    r_gm_e = _r
    ws.cell(row=_r, column=1, value=labels['gm_end'])
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('gm_end')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
    _r += 1

    r_gm = _r; L(_r, labels['fc_gm']); _r += 1
    B(r_gm, 2, h_gm, '0.0%')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_gm, col, f"={c}{r_gm_s}+({c}{r_gm_e}-{c}{r_gm_s})*{i - 1}/{N - 1}", '0.0%')

    r_gp = _r; L(_r, f"{labels['fc_gross_profit']} ({ccy_sym}{unit_suffix})"); _r += 1
    if h_rev_B and h_gm:
        B(r_gp, 2, h_rev_B * h_gm, '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_gp, col, f"={c}{r_total}*{c}{r_gm}")

    r_exp_s = _r
    ws.cell(row=_r, column=1, value=labels['exp_start'])
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('expense_start')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
    _r += 1

    r_exp_e = _r
    ws.cell(row=_r, column=1, value=labels['exp_end'])
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('expense_end')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
    _r += 1

    r_exp = _r; L(_r, labels['fc_expense_ratio']); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_exp, col, f"={c}{r_exp_s}+({c}{r_exp_e}-{c}{r_exp_s})*{i - 1}/{N - 1}", '0.0%')

    r_opex = _r; L(_r, f"{labels['fc_opex']} ({ccy_sym}{unit_suffix})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_opex, col, f"={c}{r_total}*{c}{r_exp}")

    r_ebit = _r; L(_r, f"{labels['fc_ebit']} ({ccy_sym}{unit_suffix})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_ebit, col, f"={c}{r_gp}-{c}{r_opex}")

    r_tax = _r
    ws.cell(row=_r, column=1, value=labels['tax_rate'])
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('tax_rate')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0%')
    _r += 1

    r_ni = _r; L(_r, f"{labels['fc_net_income']} ({ccy_sym}{unit_suffix})"); _r += 1
    B(r_ni, 2, h_ni_B, '#,##0.0')
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        O(r_ni, col, f"={c}{r_ebit}*(1-{c}{r_tax})")

    # ============================================
    # 估值
    # ============================================
    S(_r, labels['section_val_forecast']); _r += 1
    r_pe_s = _r
    ws.cell(row=_r, column=1, value=f"{labels['target_pe']} 起始")
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('pe_start')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0')
    _r += 1

    r_pe_e = _r
    ws.cell(row=_r, column=1, value=f"{labels['target_pe']} 终年")
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('pe_end')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='0.0')
    _r += 1

    r_pe = _r; L(_r, labels['fc_target_pe']); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_pe, col, f"={c}{r_pe_s}+({c}{r_pe_e}-{c}{r_pe_s})*{i - 1}/{N - 1}", '0.0')

    r_mkt = _r; L(_r, f"{labels['fc_implied_mkt_cap']} ({ccy_sym}{unit_suffix})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        O(r_mkt, col, f"={c}{r_ni}*{c}{r_pe}")

    r_sh = _r
    ws.cell(row=_r, column=1, value=labels['fc_shares'])
    ws.cell(row=_r, column=1).font = BLACK_BOLD; ws.cell(row=_r, column=1).border = THIN_BORDER
    for i in range(TC):
        ws.cell(row=_r, column=i + 2, value=f"={asum_co('shares_m')}")
        style_data_cell(ws, _r, i + 2, font=GREEN_FONT, num_fmt='#,##0')
    _r += 1

    r_price = _r; L(_r, f"{labels['implied_price_label']} ({ccy_sym})"); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        O(r_price, col, f"={c}{r_mkt}*1000/{c}{r_sh}", '#,##0.00')

    r_prem = _r; L(_r, labels['fc_premium']); _r += 1
    for i in range(1, N + 1):
        col = i + 2; c = get_column_letter(col)
        K(r_prem, col, f"=({c}{r_price}-{VAL})/{VAL}", '0.0%')

    ws.column_dimensions['A'].width = 28
    for c in range(2, TC + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14


def build_fy_info_sheet(ws, fy_analysis, labels):
    """财年信息工作表：详细说明 FY→自然年映射逻辑"""
    ws.title = labels['sheet_fy_info']

    r = 1
    ws.cell(row=r, column=1, value="Fiscal Year → Calendar Year Mapping")
    ws.cell(row=r, column=1).font = Font(size=14, bold=True, color="1F4E79")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    r = 3
    ws.cell(row=r, column=1, value="FY End Month:")
    ws.cell(row=r, column=1).font = BLACK_BOLD
    ws.cell(row=r, column=2, value=fy_analysis.get('fy_end_month', 'N/A'))

    r = 4
    ws.cell(row=r, column=1, value="FY End Date:")
    ws.cell(row=r, column=1).font = BLACK_BOLD
    ws.cell(row=r, column=2, value=fy_analysis.get('fiscal_year_end', 'N/A'))

    r = 5
    is_cal = fy_analysis.get('is_calendar_year', False)
    ws.cell(row=r, column=1, value="Calendar Year Aligned:")
    ws.cell(row=r, column=1).font = BLACK_BOLD
    ws.cell(row=r, column=2, value="Yes" if is_cal else "No — requires mapping")
    ws.cell(row=r, column=2).font = Font(color="FF0000" if not is_cal else "008000", bold=True)

    r = 7
    ws.cell(row=r, column=1, value="Mapping Rule:")
    ws.cell(row=r, column=1).font = BLACK_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value=fy_analysis.get('mapping_rule', ''))
    ws.cell(row=r, column=1).alignment = WRAP_ALIGN

    r = 9
    ws.cell(row=r, column=1, value="Detailed Mappings:")
    ws.cell(row=r, column=1).font = Font(size=12, bold=True, color="1F4E79")

    r = 10
    ws.cell(row=r, column=1, value="Report Date")
    ws.cell(row=r, column=2, value="FY Label")
    ws.cell(row=r, column=3, value="Natural Year")
    style_header_row(ws, r, 1, 3)

    mappings = fy_analysis.get('date_mappings', {})
    fy_end_month = fy_analysis.get('fy_end_month', 12)

    for i, (date_str, nat_year) in enumerate(sorted(mappings.items())):
        row = r + 1 + i
        if fy_end_month <= 6:
            fy_label = f"FY{nat_year + 1}"
        else:
            fy_label = f"FY{nat_year}"

        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=fy_label)
        ws.cell(row=row, column=3, value=f"{nat_year}A")
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER

    note_row = r + 1 + len(mappings) + 2
    ws.cell(row=note_row, column=1, value="Full Notes:")
    ws.cell(row=note_row, column=1).font = BLACK_BOLD
    note = fy_analysis.get('fiscal_year_note', '')
    ws.cell(row=note_row + 1, column=1, value=note)
    ws.cell(row=note_row + 1, column=1).alignment = WRAP_ALIGN
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 3, end_column=4)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16


def build_sanity_checks(ws, checks, labels):
    ws.title = labels['sheet_sanity']

    ws.cell(row=1, column=1, value=labels['label_metric'])
    ws.cell(row=1, column=2, value="Bear")
    ws.cell(row=1, column=3, value="Base")
    ws.cell(row=1, column=4, value="Bull")
    ws.cell(row=1, column=5, value="Status")
    style_header_row(ws, 1, 1, 5)

    for r_idx, check in enumerate(checks, start=2):
        ws.cell(row=r_idx, column=1, value=check['name'])
        ws.cell(row=r_idx, column=1).font = BLACK_BOLD
        ws.cell(row=r_idx, column=1).border = THIN_BORDER
        for c_idx, scenario in enumerate(['bear', 'base', 'bull'], start=2):
            ws.cell(row=r_idx, column=c_idx, value=check.get(scenario, ''))
            style_data_cell(ws, r_idx, c_idx)
        ws.cell(row=r_idx, column=5, value=check.get('status', ''))
        style_data_cell(ws, r_idx, 5)

    ws.column_dimensions['A'].width = 30
    for c in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[c].width = 14


def main():
    parser = argparse.ArgumentParser(description='生成估值模型 Excel')
    parser.add_argument('--ticker', required=True, help='股票代码')
    parser.add_argument('--input', help='模型数据 JSON 文件路径')
    parser.add_argument('--output-dir', default='./out', help='输出目录')
    parser.add_argument('--projection-years', type=int, default=10, help='预测年数')
    parser.add_argument('--language', choices=['zh-CN', 'en'], default=None,
                        help='输出语言（默认自动检测）')
    args = parser.parse_args()

    data = {}
    val_data = {}
    assumptions = {}
    segment_assumptions = {}
    checks = []
    sources = {}
    fy_analysis = {}
    language = args.language or 'en'
    ccy_sym = '$'
    unit_suffix = 'M'

    if args.input:
        with open(args.input, 'r', encoding='utf-8') as f:
            raw = json.load(f)
            data = raw.get('hist_data', raw.get('data', {}))
            val_data = raw.get('val_data', {})
            assumptions = raw.get('assumptions', {})
            segment_assumptions = raw.get('segment_assumptions', {})
            checks = raw.get('checks', [])
            sources = raw.get('sources', {})
            fy_analysis = raw.get('fiscal_year_analysis', {})
            if not args.language:
                language = raw.get('language', 'en')
            ccy_sym = raw.get('currency_symbol', '$')
            unit_suffix = raw.get('unit_suffix', 'M')

    labels = get_labels(language)

    wb = Workbook()

    ws1 = wb.active
    build_financial_history(ws1, data, labels, ccy_sym, unit_suffix)

    ws2 = wb.create_sheet()
    build_valuation_history(ws2, val_data, labels, ccy_sym)

    if fy_analysis:
        ws_fy = wb.create_sheet()
        build_fy_info_sheet(ws_fy, fy_analysis, labels)

    if assumptions:
        ws3 = wb.create_sheet()
        build_assumptions(ws3, assumptions, labels, ccy_sym, 'B', sources)

        ws4 = wb.create_sheet()
        current_price = val_data.get('current_price')
        build_projections(ws4, assumptions, data, labels, ccy_sym, 'B',
                          args.projection_years, current_price=current_price)

        if checks:
            ws5 = wb.create_sheet()
            build_sanity_checks(ws5, checks, labels)

    elif segment_assumptions:
        ws3 = wb.create_sheet()
        build_segment_assumptions(ws3, segment_assumptions, labels, ccy_sym, 'B', sources)

        ws4 = wb.create_sheet()
        current_price = val_data.get('current_price')
        build_segment_projections(ws4, segment_assumptions, data, labels, ccy_sym, 'B',
                                  args.projection_years, current_price=current_price)

        if checks:
            ws5 = wb.create_sheet()
            build_sanity_checks(ws5, checks, labels)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{args.ticker}_coverage_model.xlsx"
    wb.save(str(output_path))

    print(f"模型已生成: {output_path}")
    return str(output_path)


if __name__ == '__main__':
    main()
